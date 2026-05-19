"""Excel summary export.

Sheet 1 (Summary):
    Product info / dimensions / computed result.

Sheet 2 (Top 5 Solutions):
    The full top-N comparison table.

Sheet 3 (Layer Placements):
    Per-carton x,y,z, dimensions, barcode visibility — for QA / re-use.
"""
from __future__ import annotations

from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ..core   import compare_solutions
from ..models import StackingResult


BLUE = "1F4E8E"
LIGHT_BLUE = "E7EEF7"


def _style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(ws, cols: int):
    for c in range(1, cols + 1):
        max_len = 0
        for cell in ws[get_column_letter(c)]:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[get_column_letter(c)].width = max_len + 2


def export_excel(output_path: str,
                 primary: StackingResult,
                 top_solutions: Optional[List[StackingResult]] = None,
                 product_name: str = "—",
                 product_code: str = "—",
                 pallet_type: str = "Standard",
                 pallet_weight: float = 0.0) -> str:
    wb = Workbook()

    # ---- Sheet 1: Summary ------------------------------------------------
    ws = wb.active
    ws.title = "Summary"

    rows = [
        ("Product Name",        product_name),
        ("Product Code",        product_code),
        ("Pallet Type",         pallet_type),
        ("Case (L × W × H)",
         f"{primary.carton.length:.1f} × {primary.carton.width:.1f} "
         f"× {primary.carton.height:.1f} mm"),
        ("Case Weight (kg)",    f"{primary.carton.weight:.3f}"),
        ("Pallet (L × W)",
         f"{primary.pallet.length:.1f} × {primary.pallet.width:.1f} mm"),
        ("Pallet Height (mm)",  f"{primary.pallet.height:.1f}"),
        ("Max Total Height",    f"{primary.pallet.max_total_height:.1f}"),
        ("Reserved Margins",
         f"F={primary.pallet.margin_front} B={primary.pallet.margin_back} "
         f"L={primary.pallet.margin_left} R={primary.pallet.margin_right}"),
        ("Empty Pallet Weight", f"{pallet_weight:.3f} kg"),
        ("",                    ""),
        ("Layout",              primary.layout_name),
        ("Cases per Layer",     primary.cases_per_layer),
        ("Layers per Pallet",   primary.layer_count),
        ("Total Cases",         primary.total_cases),
        ("Total Height (mm)",   f"{primary.total_height():.1f}"),
        ("Net Weight (kg)",     f"{primary.total_cases * primary.carton.weight:.3f}"),
        ("Gross Weight (kg)",
         f"{primary.total_cases * primary.carton.weight + pallet_weight:.3f}"),
        ("Area Utilization (%)",   f"{primary.area_utilization * 100:.2f}"),
        ("Volume Utilization (%)", f"{primary.volume_utilization * 100:.2f}"),
        ("Barcode Exposure (%)",   f"{primary.barcode_exposure * 100:.2f}"),
        ("Interlock",              "Yes" if primary.interlock else "No"),
        ("Score",                  f"{primary.score:.2f}"),
    ]
    ws["A1"] = "Pallet Stacking Summary"
    ws["A1"].font = Font(bold=True, size=14, color=BLUE)
    ws.merge_cells("A1:B1")

    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        ws.cell(row=i, column=2, value=v)
    _autosize(ws, 2)

    # ---- Sheet 2: Top-N --------------------------------------------------
    ws2 = wb.create_sheet("Top 5 Solutions")
    head = ["Rank", "Layout", "Vertical axis", "Cases/Layer", "Layers",
            "Total", "Area %", "Volume %", "Barcode %", "Interlock", "Score"]
    for c, h in enumerate(head, start=1):
        cell = ws2.cell(row=1, column=c, value=h)
        _style_header(cell)
    if top_solutions:
        for ri, r in enumerate(compare_solutions(top_solutions), start=2):
            ws2.cell(row=ri, column=1, value=r["rank"])
            ws2.cell(row=ri, column=2, value=r["layout"])
            ws2.cell(row=ri, column=3, value=r["vertical_axis"])
            ws2.cell(row=ri, column=4, value=r["cases_per_layer"])
            ws2.cell(row=ri, column=5, value=r["layer_count"])
            ws2.cell(row=ri, column=6, value=r["total_cases"])
            ws2.cell(row=ri, column=7, value=r["area_util_%"])
            ws2.cell(row=ri, column=8, value=r["volume_util_%"])
            ws2.cell(row=ri, column=9, value=r["barcode_exposure_%"])
            ws2.cell(row=ri, column=10, value="Yes" if r["interlock"] else "No")
            ws2.cell(row=ri, column=11, value=r["score"])
    _autosize(ws2, len(head))

    # ---- Sheet 3: Placements --------------------------------------------
    ws3 = wb.create_sheet("Layer Placements")
    head = ["Layer", "Index", "X (mm)", "Y (mm)", "Z (mm)",
            "dX", "dY", "dZ", "Rotation", "face_X", "face_Y", "face_Z",
            "Barcode Visible"]
    for c, h in enumerate(head, start=1):
        cell = ws3.cell(row=1, column=c, value=h)
        _style_header(cell)
    row = 2
    for li, layer in enumerate(primary.layers, start=1):
        for ci, p in enumerate(layer.placements, start=1):
            ws3.cell(row=row, column=1, value=li)
            ws3.cell(row=row, column=2, value=ci)
            ws3.cell(row=row, column=3, value=round(p.x, 1))
            ws3.cell(row=row, column=4, value=round(p.y, 1))
            ws3.cell(row=row, column=5, value=round(p.z, 1))
            ws3.cell(row=row, column=6, value=round(p.dx, 1))
            ws3.cell(row=row, column=7, value=round(p.dy, 1))
            ws3.cell(row=row, column=8, value=round(p.dz, 1))
            ws3.cell(row=row, column=9, value=p.rotation)
            ws3.cell(row=row, column=10, value=p.face_x)
            ws3.cell(row=row, column=11, value=p.face_y)
            ws3.cell(row=row, column=12, value=p.face_z)
            ws3.cell(row=row, column=13,
                     value="Yes" if p.barcode_visible else "No")
            row += 1
    _autosize(ws3, len(head))

    wb.save(output_path)
    return output_path
