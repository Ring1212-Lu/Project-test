#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
棧板包裝資料處理工具 (v2.0 — code review 修正版)
需要安裝套件: pip install pandas openpyxl
打包 EXE: pyinstaller --onefile --windowed --name 棧板包裝工具 packing_tool.py
"""

import math, os, sys, threading, tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
import unicodedata
import traceback
import pandas as pd
import openpyxl

# ======================== 固定參數 ========================
LOADING_TYPE_MAP = {'空運':1,'空運一般':1,'海運一般':2,'陸運':4,'海運散貨':6}
DEFAULT_AIR_LIMIT, DEFAULT_SEA_LIMIT = 150, 220
PALLET_HEIGHT = {'EU':14.1,'L7-G':14.1,'EPAL-EU':16.5,'EPAL-STD':18.5}
PALLET_WIDTH = {'EU':80,'L7-G':100,'EPAL-EU':80,'EPAL-STD':100}
PALLET_TYPE_TO_CODE = {'EU':'EU','Standard':'L7-G'}
IGNORED_PALLET_TYPES = {'Big box','Big Box'}
PALLET_LENGTH = 120
PRODUCT_LINE_OPTIONS = ['AIO(PT)', 'PF']
LOADING_NAME = {1:'空運',2:'海運一般',4:'陸運',6:'海運散貨'}

C = {
    'bg':'#f5f7fa','white':'#ffffff',
    'blue':'#4a7fb5','blue_light':'#dce8f3','blue_pale':'#eef4fa',
    'green':'#5b9a6a','green_light':'#d6edda','green_pale':'#eef7f0',
    'orange':'#d48a45','orange_light':'#fbe8d2','orange_pale':'#fef6ed',
    'red':'#c0392b','red_light':'#f8d7da','red_pale':'#fdf0f0',
    'border':'#d0d8e0','grid':'#e2e8ee',
    'text':'#2c3e50','text_dim':'#7f8c9b','text_label':'#445566',
    'input_bg':'#ffffff','input_bd':'#c5d0db','input_focus':'#4a7fb5',
    'btn_fg':'#ffffff',
    'btn_blue':'#4a7fb5','btn_hover':'#3d6d9e',
    'btn_green':'#5b9a6a','btn_green_h':'#4d8559',
    'btn_orange':'#d48a45','btn_orange_h':'#bf7a3b',
    'prog_bg':'#e2e8ee','prog_fg':'#5b9a6a',
    'edited':'#fff3cd','edited_border':'#f0c040',
}


# ======================== 工具函數 ========================

def clean_str(raw):
    """將所有 Unicode 空白類字元 (\\xa0, \\u3000 等) 替換為普通空格，再 strip"""
    if pd.isna(raw):
        return ''
    s = str(raw)
    s = ''.join(' ' if unicodedata.category(c) == 'Zs' else c for c in s)
    return s.strip()


def has_rename_marker(raw_str):
    """檢查是否含改名標記 => (半形) 或 ＝＞ (全形)"""
    return '=>' in raw_str or '\uff1d\uff1e' in raw_str


def resolve_location(raw):
    s = clean_str(raw)
    if not s:
        return None
    for sep in ['=>', '\uff1d\uff1e']:  # 半形 => 和全形 ＝＞
        if sep in s:
            n = s.split(sep)[-1].strip()
            for p in ['改名為', '改']:
                if n.startswith(p):
                    n = n[len(p):]
                    break
            return n.strip() or None
    return s


def calc_layer(lim, pc, bh):
    if lim is None or bh is None or bh <= 0:
        return 1
    return max(1, math.floor((lim - PALLET_HEIGHT[pc]) / bh))


def get_desktop():
    if sys.platform == 'win32':
        desktop = Path(os.environ.get('USERPROFILE', '')) / 'Desktop'
    else:
        desktop = Path.home() / 'Desktop'
    desktop.mkdir(parents=True, exist_ok=True)
    return desktop


def find_product_col(hr, pl):
    t = ['PF'] if pl == 'PF' else ['AIO(PT)', 'PT\n(AIO)']
    for i, v in enumerate(hr):
        if not pd.isna(v) and str(v).strip() in t:
            return i
    return None


# ======================== 刪除線偵測 ========================

def get_strikethrough_locations(excel_path):
    """回傳所有有刪除線的原始儲存格文字 set"""
    wb = openpyxl.load_workbook(excel_path, data_only=False, read_only=True)
    strike_set = set()
    for sname in ['歐規棧板', '海空運限高', '陸運', 'EPAL-Amazon', '日字-實木. EPAL#2']:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        for row in ws.iter_rows(min_row=3, max_col=2):
            if len(row) < 2:
                continue
            cell = row[1]
            if cell.value and cell.font and cell.font.strikethrough:
                raw = clean_str(cell.value)
                if raw:
                    strike_set.add(raw)
    wb.close()
    return strike_set


# ======================== 資料處理核心 ========================

def _split_loading_codes(cd):
    """將堆疊方式 code 拆分為列表 (處理 '/' 分隔)"""
    if '/' in cd:
        return [c.strip() for c in cd.split('/')]
    return [cd]


def collect_layer_rules(excel_path, product_line, box_height):
    """
    回傳:
      rules: 層數規則列表
      loc_data: location 資料 (含 excluded_locs 供 generate_csv 使用)
      anomalies: 異常列表
    """
    strike_raw_set = get_strikethrough_locations(excel_path)
    sheets = pd.read_excel(excel_path, sheet_name=None, header=None)
    bh = float(box_height) if box_height else None
    anomalies = []

    def check_strike(raw_str):
        return raw_str in strike_raw_set

    def process_location(raw_val, sheet_name, pf_col, row_series,
                         extra_fields=None):
        raw_loc = clean_str(raw_val)
        loc = resolve_location(raw_val)
        if not loc:
            return None, None, True
        pf_val = str(row_series.iloc[pf_col]).strip() if pf_col is not None else ''
        if 'V' not in pf_val:
            return None, None, True
        ef = extra_fields or {}
        if check_strike(raw_loc):
            anomalies.append({
                'location': raw_loc, 'sheet': sheet_name,
                'pallet_type': ef.get('pallet_type', ''),
                'loading_code': ef.get('loading_code', ''),
                'limitation': ef.get('limitation', ''),
                'reason': '有刪除線(已停用)，已排除',
                'type': 'strikethrough'})
            return None, None, True
        if has_rename_marker(raw_loc):
            anomalies.append({
                'location': raw_loc, 'sheet': sheet_name,
                'pallet_type': ef.get('pallet_type', ''),
                'loading_code': ef.get('loading_code', ''),
                'limitation': ef.get('limitation', ''),
                'reason': f'含改名標記(=>)，解析為: {loc}',
                'type': 'renamed'})
        return loc, raw_loc, False

    # ── 海空運限高 ──
    sea_air_data = {}
    sa = sheets.get('海空運限高')
    if sa is not None:
        pf = find_product_col(sa.iloc[1].tolist(), product_line)
        if pf is not None:
            for i in range(2, len(sa)):
                r = sa.iloc[i]
                pt = clean_str(r.iloc[9])
                cd = clean_str(r.iloc[11])
                lim_raw = r.iloc[10]
                lim_str = clean_str(lim_raw) if not pd.isna(lim_raw) else ''

                loc, raw_loc, skip = process_location(
                    r.iloc[1], '海空運限高', pf, r,
                    {'pallet_type': pt, 'loading_code': cd, 'limitation': lim_str})
                if skip:
                    continue

                if pt in IGNORED_PALLET_TYPES:
                    anomalies.append({'location': loc, 'sheet': '海空運限高',
                        'pallet_type': pt, 'loading_code': cd, 'limitation': lim_str,
                        'reason': f'Pallet type={pt}，已設定忽略', 'type': 'ignored'})
                    continue
                if pt and pt not in PALLET_TYPE_TO_CODE:
                    anomalies.append({'location': loc, 'sheet': '海空運限高',
                        'pallet_type': pt, 'loading_code': cd, 'limitation': lim_str,
                        'reason': f'未知的 Pallet type="{pt}"', 'type': 'unknown'})
                if cd:
                    for sc in _split_loading_codes(cd):
                        if sc and sc not in LOADING_TYPE_MAP:
                            anomalies.append({'location': loc, 'sheet': '海空運限高',
                                'pallet_type': pt, 'loading_code': cd, 'limitation': lim_str,
                                'reason': f'未知的堆疊方式Code="{sc}"', 'type': 'unknown'})
                try:
                    lv = float(lim_raw)
                except (ValueError, TypeError):
                    lv = None
                    if lim_str:
                        anomalies.append({'location': loc, 'sheet': '海空運限高',
                            'pallet_type': pt, 'loading_code': cd, 'limitation': lim_str,
                            'reason': f'限高無法解析: "{lim_str}"', 'type': 'parse_error'})

                pc = PALLET_TYPE_TO_CODE.get(pt, 'EU')
                sea_air_data.setdefault(loc, []).append((cd, pc, lv))

    # ── 陸運 ──
    land_data, land_lim = {}, {}
    ls = sheets.get('陸運')
    if ls is not None:
        pf = find_product_col(ls.iloc[1].tolist(), product_line)
        if pf is not None:
            for i in range(2, len(ls)):
                r = ls.iloc[i]
                pt = clean_str(r.iloc[8])

                loc, raw_loc, skip = process_location(
                    r.iloc[1], '陸運', pf, r,
                    {'pallet_type': pt, 'loading_code': '陸運'})
                if skip:
                    continue

                if pt in IGNORED_PALLET_TYPES:
                    anomalies.append({'location': loc, 'sheet': '陸運',
                        'pallet_type': pt, 'loading_code': '陸運', 'limitation': '',
                        'reason': f'Pallet type={pt}，已設定忽略', 'type': 'ignored'})
                    continue
                if pt and pt not in PALLET_TYPE_TO_CODE:
                    anomalies.append({'location': loc, 'sheet': '陸運',
                        'pallet_type': pt, 'loading_code': '陸運', 'limitation': '',
                        'reason': f'未知的 Pallet type="{pt}"', 'type': 'unknown'})

                pc = PALLET_TYPE_TO_CODE.get(pt, 'EU')
                land_data.setdefault(loc, []).append(pc)
                try:
                    land_lim[loc] = float(r.iloc[9])
                except (ValueError, TypeError):
                    pass

    # ── 歐規棧板 ──
    eu_locations = set()
    es = sheets.get('歐規棧板')
    if es is not None:
        pf = find_product_col(es.iloc[1].tolist(), product_line)
        if pf is not None:
            for i in range(2, len(es)):
                r = es.iloc[i]
                loc, raw_loc, skip = process_location(
                    r.iloc[1], '歐規棧板', pf, r, {'pallet_type': 'EU'})
                if skip:
                    continue
                eu_locations.add(loc)

    # ── EPAL-Amazon ──
    epal_eu_locs = set()
    ep = sheets.get('EPAL-Amazon')
    if ep is not None:
        pf = find_product_col(ep.iloc[1].tolist(), product_line)
        if pf is not None:
            for i in range(2, len(ep)):
                r = ep.iloc[i]
                loc, raw_loc, skip = process_location(
                    r.iloc[1], 'EPAL-Amazon', pf, r, {'pallet_type': 'EPAL-EU'})
                if skip:
                    continue
                epal_eu_locs.add(loc)

    # ── 日字-實木 EPAL#2 ──
    epal_std_locs = set()
    ep2 = sheets.get('日字-實木. EPAL#2')
    if ep2 is not None:
        pf = find_product_col(ep2.iloc[1].tolist(), product_line)
        if pf is not None:
            for i in range(2, len(ep2)):
                r = ep2.iloc[i]
                loc, raw_loc, skip = process_location(
                    r.iloc[1], '日字-實木 EPAL#2', pf, r, {'pallet_type': 'EPAL-STD'})
                if skip:
                    continue
                epal_std_locs.add(loc)

    # ── 彙整規則 ──
    rule_map = {}

    def add_rule(pc, lt, lim, source):
        actual_lim = float(lim) if lim is not None else _default_limit(lt)
        key = (pc, lt, actual_lim)
        if key not in rule_map:
            rule_map[key] = {
                'pallet_code': pc, 'loading_type': lt,
                'loading_name': LOADING_NAME.get(lt, str(lt)),
                'limitation': actual_lim, 'is_default': (lim is None),
                'calc_layer': calc_layer(actual_lim, pc, bh), 'source': source}

    for loc in eu_locations:
        e = sea_air_data.get(loc, [])
        al = [x[2] for x in e if '空運' in x[0] and x[1] == 'EU' and x[2] is not None]
        for a in al:
            add_rule('EU', 1, a, '歐規棧板')
        if not al:
            add_rule('EU', 1, None, '歐規棧板')
        sl = [x[2] for x in e if '海運' in x[0] and '散貨' not in x[0]
              and x[1] == 'EU' and x[2] is not None]
        for s in sl:
            add_rule('EU', 2, s, '海空運限高')
        if not sl:
            add_rule('EU', 2, None, '歐規棧板(預設)')

    for loc, entries in sea_air_data.items():
        for cd, pc, lv in entries:
            if pc == 'EU' and loc in eu_locations:
                if cd == '海運散貨':
                    add_rule('EU', 6, lv, '海空運限高')
                continue
            for sc in _split_loading_codes(cd):
                lt = LOADING_TYPE_MAP.get(sc)
                if lt is not None:
                    add_rule(pc, lt, lv, '海空運限高')

    for loc, pcs in land_data.items():
        for pc in pcs:
            add_rule(pc, 4, land_lim.get(loc), '陸運')

    if epal_eu_locs:
        add_rule('EPAL-EU', 1, 175, 'EPAL-Amazon')
        add_rule('EPAL-EU', 2, 175, 'EPAL-Amazon')
    if epal_std_locs:
        add_rule('EPAL-STD', 1, 180, '日字EPAL#2')
        add_rule('EPAL-STD', 2, 180, '日字EPAL#2')
    add_rule('EU', 1, None, 'ASTPHQ(固定)')
    add_rule('EU', 2, None, 'ASTPHQ(固定)')

    rules = sorted(rule_map.values(),
                   key=lambda x: (x['pallet_code'], x['loading_type'], x['limitation']))

    loc_data = {
        'sea_air_data': sea_air_data, 'land_data': land_data,
        'land_lim': land_lim, 'eu_locations': eu_locations,
        'epal_eu_locs': epal_eu_locs, 'epal_std_locs': epal_std_locs,
    }
    return rules, loc_data, anomalies


def _default_limit(lt):
    return float(DEFAULT_AIR_LIMIT if lt == 1 else DEFAULT_SEA_LIMIT)


def generate_csv(carton_code, box_height, bl1, bl2, loc_data,
                 layer_lookup, excluded_locations=None):
    """
    產生 CSV。
    excluded_locations: set of location 名稱，這些會被排除不產出。
    """
    hv = float(box_height) if box_height else None
    sad = loc_data['sea_air_data']
    ld, ll = loc_data['land_data'], loc_data['land_lim']
    eu_set = loc_data['eu_locations']
    excl = excluded_locations or set()
    rows = []

    def gl(pc, lt, lim):
        key = (pc, lt, float(lim) if lim is not None else _default_limit(lt))
        if key in layer_lookup:
            return layer_lookup[key]
        if lim is not None and hv and hv > 0:
            return calc_layer(lim, pc, hv)
        return calc_layer(_default_limit(lt), pc, hv)

    def mr(pc, loc, lt, bl, layer):
        return {
            'Carton Code': carton_code, 'Length': PALLET_LENGTH,
            'Width': PALLET_WIDTH[pc], 'Hight': hv,
            'Ship to locaton': loc, 'Pallet Code': pc,
            'Loading Type': lt, 'Bottom Loading': bl,
            'Bottom Layer': layer, 'Bottom Place Type': 'H',
            'Top Loading': None, 'Top Layer': None, 'Top Place Type': None,
            'Top Cover': None, 'Side Cover': None,
            'Corner Protector Weight': None,
            'MOQ': layer if pc == 'EPAL-EU' else None}

    # 歐規棧板
    for loc in sorted(eu_set):
        if loc in excl:
            continue
        e = sad.get(loc, [])
        al = [x[2] for x in e if '空運' in x[0] and x[1] == 'EU' and x[2] is not None]
        sl = [x[2] for x in e if '海運' in x[0] and '散貨' not in x[0]
              and x[1] == 'EU' and x[2] is not None]
        rows.append(mr('EU', loc, 1, bl1, gl('EU', 1, min(al) if al else None)))
        rows.append(mr('EU', loc, 2, bl1, gl('EU', 2, min(sl) if sl else None)))

    # 海空運限高
    for loc in sorted(sad.keys()):
        if loc in excl:
            continue
        for cd, pc, lv in sad[loc]:
            if pc == 'EU' and loc in eu_set:
                if cd == '海運散貨':
                    rows.append(mr('EU', loc, 6, bl1, gl('EU', 6, lv)))
                continue
            bl = bl1 if pc in ('EU', 'EPAL-EU') else bl2
            for sc in _split_loading_codes(cd):
                lt = LOADING_TYPE_MAP.get(sc)
                if lt is not None:
                    rows.append(mr(pc, loc, lt, bl, gl(pc, lt, lv)))

    # 陸運
    for loc in sorted(ld.keys()):
        if loc in excl:
            continue
        for pc in ld[loc]:
            bl = bl1 if pc in ('EU', 'EPAL-EU') else bl2
            rows.append(mr(pc, loc, 4, bl, gl(pc, 4, ll.get(loc))))

    # EPAL-Amazon
    for loc in sorted(loc_data['epal_eu_locs']):
        if loc in excl:
            continue
        rows.append(mr('EPAL-EU', loc, 1, bl1, gl('EPAL-EU', 1, 175)))
        rows.append(mr('EPAL-EU', loc, 2, bl1, gl('EPAL-EU', 2, 175)))

    # EPAL-STD
    for loc in sorted(loc_data['epal_std_locs']):
        if loc in excl:
            continue
        rows.append(mr('EPAL-STD', loc, 1, bl2, gl('EPAL-STD', 1, 180)))
        rows.append(mr('EPAL-STD', loc, 2, bl2, gl('EPAL-STD', 2, 180)))

    # ASTPHQ
    if 'ASTPHQ' not in excl:
        rows.append(mr('EU', 'ASTPHQ', 1, bl1, gl('EU', 1, None)))
        rows.append(mr('EU', 'ASTPHQ', 2, bl1, gl('EU', 2, None)))

    df = pd.DataFrame(rows)
    co = ['Carton Code', 'Length', 'Width', 'Hight', 'Ship to locaton',
          'Pallet Code', 'Loading Type', 'Bottom Loading', 'Bottom Layer',
          'Bottom Place Type', 'Top Loading', 'Top Layer', 'Top Place Type',
          'Top Cover', 'Side Cover', 'Corner Protector Weight', 'MOQ']
    df = df[co].drop_duplicates(
        subset=['Ship to locaton', 'Pallet Code', 'Loading Type'], keep='first')
    df = df.sort_values(
        ['Pallet Code', 'Ship to locaton', 'Loading Type']).reset_index(drop=True)

    op = get_desktop() / 'C.CSV'
    from datetime import datetime
    ts = datetime.now().strftime('%H%M%S')
    backup = get_desktop() / f'C_backup_{ts}.CSV'
    try:
        op.rename(backup)
    except OSError:
        pass  # file doesn't exist or can't be renamed — proceed to write
    df.to_csv(op, index=False, encoding='utf-8-sig')
    return str(op), len(df)


# ======================== GUI ========================

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("棧板包裝資料處理工具 v2.0")
        self.root.geometry("880x920")
        self.root.minsize(880, 700)
        self.root.resizable(True, True)
        self.root.configure(bg=C['bg'])
        self.rules = []
        self.loc_data = None
        self.anomalies = []
        self.layer_entries = {}
        self.excluded_anomalies = {}
        self._running = False
        self._build()

    def _section(self, parent, title, ck='blue'):
        color, light, pale = C[ck], C[f'{ck}_light'], C[f'{ck}_pale']
        outer = tk.Frame(parent, bg=C['bg'])
        outer.pack(fill='x', padx=12, pady=(0, 4))
        hdr = tk.Frame(outer, bg=pale)
        hdr.pack(fill='x')
        tk.Frame(hdr, bg=color, width=4).pack(side='left', fill='y')
        tk.Label(hdr, text=f"  {title}", font=('Segoe UI', 10, 'bold'),
                 bg=pale, fg=color, pady=4).pack(side='left')
        bw = tk.Frame(outer, bg=C['border'])
        bw.pack(fill='both', expand=True)
        body = tk.Frame(bw, bg=C['white'])
        body.pack(fill='both', expand=True, padx=1, pady=(0, 1))
        inner = tk.Frame(body, bg=C['white'])
        inner.pack(fill='both', expand=True)
        tk.Frame(inner, bg=light, width=3).pack(side='left', fill='y')
        content = tk.Frame(inner, bg=C['white'], padx=8, pady=4)
        content.pack(side='left', fill='both', expand=True)
        return content

    def _field_row(self, parent, label, default='', combo=None, idx=0):
        bg = C['white'] if idx % 2 == 0 else '#f8fafb'
        row = tk.Frame(parent, bg=bg)
        row.pack(fill='x')
        lf = tk.Frame(row, bg=bg, width=220)
        lf.pack(side='left', fill='y')
        lf.pack_propagate(False)
        tk.Label(lf, text=label, font=('Segoe UI', 9), bg=bg,
                 fg=C['text_label'], anchor='e', padx=8, pady=4
                 ).pack(fill='both', expand=True)
        tk.Frame(row, bg=C['grid'], width=1).pack(side='left', fill='y')
        vf = tk.Frame(row, bg=bg, padx=8, pady=2)
        vf.pack(side='left', fill='both', expand=True)
        var = tk.StringVar(value=default)
        if combo:
            ttk.Combobox(vf, textvariable=var, values=combo, state='readonly',
                         width=26, font=('Segoe UI', 10)).pack(anchor='w')
        else:
            tk.Entry(vf, textvariable=var, width=28, font=('Segoe UI', 10),
                     bg=C['input_bg'], fg=C['text'], insertbackground=C['blue'],
                     relief='solid', bd=1, highlightthickness=2,
                     highlightcolor=C['input_focus'],
                     highlightbackground=C['input_bd']).pack(anchor='w')
        tk.Frame(parent, bg=C['grid'], height=1).pack(fill='x')
        return var

    def _pill(self, parent, text, ck='blue'):
        f = tk.Frame(parent, bg=C[f'{ck}_light'], padx=6, pady=1)
        f.pack(side='left', padx=(0, 4))
        tk.Label(f, text=text, font=('Segoe UI', 7),
                 bg=C[f'{ck}_light'], fg=C[ck]).pack()

    def _ref_row(self, parent, pairs, idx=0):
        bg = C['white'] if idx % 2 == 0 else '#f8fafb'
        row = tk.Frame(parent, bg=bg)
        row.pack(fill='x')
        for i, (k, v) in enumerate(pairs):
            if i > 0:
                tk.Frame(row, bg=C['grid'], width=1).pack(side='left', fill='y')
            cell = tk.Frame(row, bg=bg, padx=6, pady=2)
            cell.pack(side='left', fill='both', expand=True)
            if k:
                tk.Label(cell, text=k, font=('Segoe UI', 7),
                         bg=bg, fg=C['text_dim']).pack(anchor='w')
            tk.Label(cell, text=v, font=('Segoe UI', 8, 'bold'),
                     bg=bg, fg=C['text']).pack(anchor='w')
        tk.Frame(parent, bg=C['grid'], height=1).pack(fill='x')

    def _readonly_entry(self, parent, text, width=20, bg='#ffffff'):
        e = tk.Entry(parent, width=width, font=('Segoe UI', 9),
                     bg=bg, fg=C['text'], relief='flat', bd=0,
                     readonlybackground=bg, highlightthickness=0)
        e.insert(0, text)
        e.config(state='readonly')
        return e

    def _bind_mousewheel(self, canvas):
        def _on_wheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        def _on_linux_up(event):
            canvas.yview_scroll(-1, "units")
        def _on_linux_down(event):
            canvas.yview_scroll(1, "units")

        if sys.platform in ('win32', 'darwin'):
            canvas.bind_all('<MouseWheel>', _on_wheel)
        else:
            canvas.bind_all('<Button-4>', _on_linux_up)
            canvas.bind_all('<Button-5>', _on_linux_down)

    def _build(self):
        top = tk.Frame(self.root, bg=C['white'], padx=14, pady=8)
        top.pack(fill='x')
        tk.Frame(top, bg=C['blue'], width=5, height=24).pack(side='left', padx=(0, 8))
        tf = tk.Frame(top, bg=C['white'])
        tf.pack(side='left')
        tk.Label(tf, text="棧板包裝資料處理工具", font=('Segoe UI', 13, 'bold'),
                 bg=C['white'], fg=C['text']).pack(anchor='w')
        tk.Label(tf, text="Step1: 計算預覽+異常偵測  →  Step2: 確認產出",
                 font=('Segoe UI', 8), bg=C['white'], fg=C['text_dim']).pack(anchor='w')
        flow = tk.Frame(top, bg=C['white'])
        flow.pack(side='right')
        for txt, ck in [("📄Excel", 'blue'), ("→", ""), ("🔍預覽", 'green'),
                         ("→", ""), ("⚠️檢查", 'orange'), ("→", ""), ("📊CSV", 'green')]:
            if txt == "→":
                tk.Label(flow, text="→", font=('Segoe UI', 8),
                         bg=C['white'], fg=C['text_dim']).pack(side='left', padx=2)
            else:
                self._pill(flow, txt, ck)
        tk.Frame(self.root, bg=C['border'], height=1).pack(fill='x')

        container = tk.Frame(self.root, bg=C['bg'])
        container.pack(fill='both', expand=True)
        self.canvas = tk.Canvas(container, bg=C['bg'], highlightthickness=0)
        vsb = ttk.Scrollbar(container, orient='vertical', command=self.canvas.yview)
        self.main = tk.Frame(self.canvas, bg=C['bg'])
        self.main.bind('<Configure>',
                       lambda e: self.canvas.configure(scrollregion=self.canvas.bbox('all')))
        self.canvas.create_window((0, 0), window=self.main, anchor='nw', tags='main_win')
        self.canvas.configure(yscrollcommand=vsb.set)

        def on_resize(e):
            self.canvas.itemconfig('main_win', width=e.width)
        self.canvas.bind('<Configure>', on_resize)
        self._bind_mousewheel(self.canvas)
        self.canvas.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')

        fc = self._section(self.main, "📁  來源檔案", 'blue')
        frow = tk.Frame(fc, bg=C['white'])
        frow.pack(fill='x', pady=2)
        tk.Label(frow, text="Excel", font=('Segoe UI', 9),
                 bg=C['white'], fg=C['text_label']).pack(side='left')
        self.file_var = tk.StringVar()
        tk.Entry(frow, textvariable=self.file_var, width=42, font=('Segoe UI', 9),
                 bg=C['input_bg'], fg=C['text'], insertbackground=C['blue'],
                 relief='solid', bd=1, highlightthickness=2,
                 highlightcolor=C['input_focus'],
                 highlightbackground=C['input_bd']).pack(side='left', padx=6)
        tk.Button(frow, text="📂 瀏覽", font=('Segoe UI', 9),
                  bg=C['btn_blue'], fg=C['btn_fg'],
                  activebackground=C['btn_hover'],
                  relief='flat', padx=10, pady=2, cursor='hand2',
                  command=self._browse).pack(side='left')

        ic = self._section(self.main, "✏️  變動資訊  Input Data", 'green')
        self.product_var = self._field_row(ic, "產品線  Product Line", 'PF',
                                           PRODUCT_LINE_OPTIONS, 0)
        self.carton_var = self._field_row(ic, "Carton Code", 'PT029', idx=1)
        self.height_var = self._field_row(ic, "箱高 (cm)  Box Height", '', idx=2)
        self.bl_eu_var = self._field_row(ic, "Bottom Loading  EU / EPAL-EU", '2', idx=3)
        self.bl_l7g_var = self._field_row(ic, "Bottom Loading  L7-G / EPAL-STD", '3', idx=4)

        rc = self._section(self.main, "📋  固定參數  Reference", 'orange')
        self._ref_row(rc, [('Loading Type', '空運:1  海運:2  陸運:4  散貨:6')], 0)
        self._ref_row(rc, [('棧板高(cm)',
                            'EU/L7-G:14.1  EPAL-EU:16.5  EPAL-STD:18.5')], 1)
        self._ref_row(rc, [('預設限高', '空運:150cm  海運:220cm'),
                           ('公式', 'floor((限高−棧板高)÷箱高)')], 2)
        self._ref_row(rc, [('固定加入', 'ASTPHQ → EU 空運+海運'),
                           ('EPAL-EU', 'MOQ = Bottom Layer')], 3)

        b1f = tk.Frame(self.main, bg=C['bg'])
        b1f.pack(pady=(6, 4))
        self.calc_btn = tk.Button(
            b1f, text="🔍  Step 1 : 計算預覽 + 異常偵測",
            font=('Segoe UI', 11, 'bold'), bg=C['btn_blue'], fg=C['btn_fg'],
            activebackground=C['btn_hover'], relief='flat', padx=24, pady=8,
            cursor='hand2', command=self._step1)
        self.calc_btn.pack()

        self.preview_frame = tk.Frame(self.main, bg=C['bg'])
        self.step2_frame = tk.Frame(self.main, bg=C['bg'])

        sf = tk.Frame(self.root, bg=C['bg'], padx=14)
        sf.pack(fill='x', side='bottom')
        self.prog_canvas = tk.Canvas(sf, height=8, bg=C['prog_bg'],
                                     highlightthickness=1,
                                     highlightbackground=C['border'])
        self.prog_canvas.pack(fill='x')
        self.status_var = tk.StringVar(value="就緒 — 請填入資訊後按 Step 1")
        tk.Label(sf, textvariable=self.status_var, font=('Segoe UI', 8),
                 bg=C['bg'], fg=C['text_dim']).pack(anchor='w', pady=(2, 4))

    def _validate(self):
        path = self.file_var.get().strip()
        if not path or not os.path.isfile(path):
            messagebox.showerror("錯誤", "請選擇有效的 Excel 檔案！")
            return None
        cc = self.carton_var.get().strip()
        if not cc:
            messagebox.showerror("錯誤", "請輸入 Carton Code！")
            return None
        h = self.height_var.get().strip()
        if not h:
            messagebox.showerror("錯誤", "請輸入箱高！")
            return None
        try:
            float(h)
        except (ValueError, TypeError):
            messagebox.showerror("錯誤", "箱高必須為數字！")
            return None
        try:
            b1 = int(self.bl_eu_var.get().strip())
            b2 = int(self.bl_l7g_var.get().strip())
        except (ValueError, TypeError):
            messagebox.showerror("錯誤", "Bottom Loading 必須為整數！")
            return None
        return path, self.product_var.get(), cc, h, b1, b2

    def _step1(self):
        if self._running:
            return
        v = self._validate()
        if not v:
            return
        self._running = True
        self.calc_btn.config(state='disabled', bg=C['border'])
        self.status_var.set("計算中 (含刪除線偵測)...")
        path, pl, cc, h, b1, b2 = v

        def work():
            try:
                result = collect_layer_rules(path, pl, h)
                self.root.after(0, lambda r=result: self._show_preview(*r))
            except Exception as e:
                err = traceback.format_exc()
                self.root.after(0, lambda: messagebox.showerror("錯誤", f"{e}\n\n{err}"))
            finally:
                self.root.after(0, self._reset_step1_btn)

        threading.Thread(target=work, daemon=True).start()

    def _reset_step1_btn(self):
        self._running = False
        self.calc_btn.config(state='normal', bg=C['btn_blue'])

    def _show_preview(self, rules, loc_data, anomalies):
        self.rules = rules
        self.loc_data = loc_data
        self.anomalies = anomalies
        self.layer_entries = {}
        self.excluded_anomalies = {}

        for w in self.preview_frame.winfo_children():
            w.destroy()
        for w in self.step2_frame.winfo_children():
            w.destroy()
        self.preview_frame.pack_forget()
        self.step2_frame.pack_forget()
        self.preview_frame.pack(fill='x')

        # ── 異常偵測 ──
        if anomalies:
            ac = self._section(self.preview_frame,
                               f"⚠️  異常偵測  ({len(anomalies)} 筆，請確認是否排除)", 'red')
            ah = tk.Frame(ac, bg=C['red'])
            ah.pack(fill='x')
            for txt, w in [("排除", 5), ("Location", 28), ("Sheet", 10), ("Pallet", 8),
                           ("Loading", 8), ("限高", 5), ("原因 Reason", 35)]:
                tk.Label(ah, text=txt, font=('Segoe UI', 8, 'bold'), bg=C['red'],
                         fg=C['btn_fg'], anchor='w', padx=4, pady=3,
                         width=w).pack(side='left')

            for i, a in enumerate(anomalies):
                bg = C['white'] if i % 2 == 0 else '#f8fafb'
                row = tk.Frame(ac, bg=bg)
                row.pack(fill='x')
                checked = a['type'] in ('ignored', 'strikethrough')
                var = tk.BooleanVar(value=checked)
                self.excluded_anomalies[i] = var
                tk.Checkbutton(row, variable=var, bg=bg,
                               activebackground=bg, width=3).pack(side='left')
                e = self._readonly_entry(row, a['location'], width=28, bg=bg)
                e.pack(side='left', padx=(0, 2))
                tk.Label(row, text=a['sheet'], font=('Segoe UI', 8), bg=bg,
                         fg=C['text_dim'], width=10, anchor='w').pack(side='left')
                tk.Label(row, text=a['pallet_type'], font=('Segoe UI', 8), bg=bg,
                         fg=C['text'], width=8, anchor='w').pack(side='left')
                tk.Label(row, text=a['loading_code'], font=('Segoe UI', 8), bg=bg,
                         fg=C['text'], width=8, anchor='w').pack(side='left')
                tk.Label(row, text=a['limitation'], font=('Segoe UI', 8), bg=bg,
                         fg=C['text'], width=5, anchor='w').pack(side='left')
                if a['type'] == 'renamed':
                    rfg = C['blue']
                elif a['type'] in ('unknown', 'parse_error'):
                    rfg = C['red']
                else:
                    rfg = C['orange']
                tk.Label(row, text=a['reason'], font=('Segoe UI', 8), bg=bg,
                         fg=rfg, anchor='w', padx=4
                         ).pack(side='left', fill='x', expand=True)
                tk.Frame(ac, bg=C['grid'], height=1).pack(fill='x')

        # ── 層數預覽 ──
        pc = self._section(self.preview_frame,
                           f"📊  層數預覽  ({len(rules)} 組規則，可直接修改層數)", 'orange')
        hdr = tk.Frame(pc, bg=C['blue'])
        hdr.pack(fill='x')
        for txt, w in [("棧板 Pallet", 14), ("運輸 Loading", 12), ("限高(cm)", 12),
                       ("來源 Source", 16), ("層數 Layer ✏️", 10)]:
            tk.Label(hdr, text=txt, font=('Segoe UI', 8, 'bold'), bg=C['blue'],
                     fg=C['btn_fg'], width=w, pady=3, anchor='w', padx=4
                     ).pack(side='left')

        for i, rule in enumerate(rules):
            bg = C['white'] if i % 2 == 0 else '#f8fafb'
            row = tk.Frame(pc, bg=bg)
            row.pack(fill='x')
            tk.Label(row, text=rule['pallet_code'], font=('Segoe UI', 9, 'bold'),
                     bg=bg, fg=C['text'], width=14, anchor='w', padx=6
                     ).pack(side='left')
            tk.Label(row, text=rule['loading_name'], font=('Segoe UI', 9),
                     bg=bg, fg=C['text'], width=12, anchor='w').pack(side='left')
            lt = f"{rule['limitation']:.0f}"
            if rule['is_default']:
                lt += " (預設)"
            tk.Label(row, text=lt, font=('Segoe UI', 9), bg=bg,
                     fg=C['text_dim'] if rule['is_default'] else C['text'],
                     width=12, anchor='w').pack(side='left')
            tk.Label(row, text=rule['source'], font=('Segoe UI', 8), bg=bg,
                     fg=C['text_dim'], width=16, anchor='w').pack(side='left')

            key = (rule['pallet_code'], rule['loading_type'], rule['limitation'])
            entry = tk.Entry(row, width=6, font=('Segoe UI', 10, 'bold'),
                             bg=C['input_bg'], fg=C['green'], justify='center',
                             relief='solid', bd=1, highlightthickness=2,
                             highlightcolor=C['orange'],
                             highlightbackground=C['input_bd'])
            entry.insert(0, str(rule['calc_layer']))
            entry.pack(side='left', padx=(4, 8), pady=2)
            self.layer_entries[key] = entry

            orig = str(rule['calc_layer'])

            def mkchk(e, o):
                def chk(*a):
                    e.config(
                        bg=C['edited'] if e.get() != o else C['input_bg'],
                        highlightbackground=(C['edited_border'] if e.get() != o
                                             else C['input_bd']))
                return chk
            entry.bind('<KeyRelease>', mkchk(entry, orig))
            tk.Frame(pc, bg=C['grid'], height=1).pack(fill='x')

        # ── Step 2 按鈕 ──
        self.step2_frame.pack(pady=(8, 16))
        tk.Label(self.step2_frame,
                 text="⬆️ 確認上方層數 + 異常排除後，按此產出 CSV ⬇️",
                 font=('Segoe UI', 9), bg=C['bg'], fg=C['text_dim']).pack(pady=(0, 4))
        bf = tk.Frame(self.step2_frame, bg=C['bg'])
        bf.pack()
        self.gen_btn = tk.Button(
            bf, text="✅  Step 2 : 確認產出  Generate C.CSV",
            font=('Segoe UI', 12, 'bold'), bg=C['btn_green'], fg=C['btn_fg'],
            activebackground=C['btn_green_h'], relief='flat', padx=24, pady=10,
            cursor='hand2', command=self._step2)
        self.gen_btn.pack(side='left', padx=4)
        tk.Button(bf, text="↺ 重新計算", font=('Segoe UI', 10),
                  bg=C['btn_orange'], fg=C['btn_fg'],
                  activebackground=C['btn_orange_h'],
                  relief='flat', padx=14, pady=10, cursor='hand2',
                  command=self._step1).pack(side='left', padx=4)

        strike_n = sum(1 for a in anomalies if a['type'] == 'strikethrough')
        rename_n = sum(1 for a in anomalies if a['type'] == 'renamed')
        amsg = ""
        if anomalies:
            parts = [f"{len(anomalies)} 筆異常"]
            if strike_n:
                parts.append(f"{strike_n} 刪除線")
            if rename_n:
                parts.append(f"{rename_n} 改名")
            amsg = f"，⚠️ {'、'.join(parts)}"
        self.status_var.set(f"✅ 預覽完成 — {len(rules)} 組規則{amsg}")
        self.root.update_idletasks()
        self.canvas.yview_moveto(0.3)

    def _step2(self):
        v = self._validate()
        if not v:
            return
        path, pl, cc, h, b1, b2 = v

        layer_lookup = {}
        for key, entry in self.layer_entries.items():
            try:
                layer_lookup[key] = int(entry.get())
            except (ValueError, TypeError):
                pc, lt, lim = key
                messagebox.showerror(
                    "錯誤",
                    f"層數必須為整數！\n{pc} / {LOADING_NAME.get(lt, lt)} / 限高{lim}")
                entry.focus_set()
                return

        excluded_locs = set()
        for idx, var in self.excluded_anomalies.items():
            if var.get():
                a = self.anomalies[idx]
                loc_name = a['location']
                resolved = resolve_location(loc_name)
                if resolved:
                    excluded_locs.add(resolved)
                cleaned = clean_str(loc_name)
                if cleaned:
                    excluded_locs.add(cleaned)

        self.gen_btn.config(state='disabled', bg=C['border'])
        self.status_var.set("產出中...")

        def work():
            try:
                out, cnt = generate_csv(
                    cc, h, b1, b2, self.loc_data, layer_lookup, excluded_locs)
                self.root.after(0, lambda: self._uprog(100, ""))
                self.root.after(0, lambda: messagebox.showinfo(
                    "完成",
                    f"✅ 共產生 {cnt} 筆資料\n"
                    f"排除 {len(excluded_locs)} 個異常 location\n\n"
                    f"輸出：{out}"))
                self.root.after(0, lambda: self.status_var.set(
                    f"✅ 已輸出 {cnt} 筆至桌面 C.CSV"))
            except Exception as e:
                err = traceback.format_exc()
                self.root.after(0, lambda: messagebox.showerror(
                    "錯誤", f"{e}\n\n{err}"))
            finally:
                self.root.after(0, lambda: self.gen_btn.config(
                    state='normal', bg=C['btn_green']))

        threading.Thread(target=work, daemon=True).start()

    def _browse(self):
        p = filedialog.askopenfilename(
            title="選擇 Excel",
            filetypes=[("Excel", "*.xlsx *.xls"), ("All", "*.*")])
        if p:
            self.file_var.set(p)

    def _uprog(self, pct, msg):
        if msg:
            self.status_var.set(msg)
        self.prog_canvas.update_idletasks()
        w = self.prog_canvas.winfo_width()
        self.prog_canvas.delete('prog')
        if pct > 0:
            self.prog_canvas.create_rectangle(
                0, 0, int(w * pct / 100), 8,
                fill=C['prog_fg'], outline='', tags='prog')
        self.root.update_idletasks()


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == '__main__':
    main()
