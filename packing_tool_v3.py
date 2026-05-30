#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
棧板包裝資料處理工具 v3.0
- 批次新增 CC（新/舊混合）
- R 版本 Excel 生成
- 特殊棧板 CSV 生成
需要套件: pip install pandas openpyxl
"""

import math, os, sys, threading, tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
import unicodedata
import traceback
from datetime import datetime
from copy import copy
import pandas as pd
import openpyxl

# ======================== Windows DPI ========================
if sys.platform == 'win32':
    try:
        import ctypes
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass

# ======================== 固定參數 ========================
LOADING_TYPE_MAP = {'空運':1,'空運一般':1,'海運一般':2,'陸運':4,'海運散貨':6}
DEFAULT_AIR_LIMIT, DEFAULT_SEA_LIMIT = 150, 220
PALLET_HEIGHT = {'EU':14.1,'L7-G':14.1,'EPAL-EU':16.5,'EPAL-STD':18.5}
PALLET_WIDTH = {'EU':80,'L7-G':100,'EPAL-EU':80,'EPAL-STD':100}
PALLET_TYPE_TO_CODE = {'EU':'EU','Standard':'L7-G'}
IGNORED_PALLET_TYPES = {'Big box','Big Box'}
PALLET_LENGTH = 120
PRODUCT_LINE_OPTIONS = ['AIO(PT)', 'PF']
LOADING_NAME = {1:'空運',2:'海運一般',4:'陸運',6:'海運散貨'}

PRODUCT_LINE_CONFIG = {
    'AIO(PT)': {
        'sheet1': 'ASUS AiO PC Info-PT',
        'sheet2': 'Freight simulation(PT)   ',
        'rev_cell': 'V2',
        'date_cell': 'S2',
        's2_rev_cell': 'P2',
        's2_date_cell': 'N2',
        'data_start': 6,
        'file_prefix': 'PT',
    },
    'PF': {
        'sheet1': 'ASUS Commercial PC Info-PF ',
        'sheet2': 'Freight simulation(PF) ',
        'rev_cell': 'W2',
        'date_cell': 'S2',
        's2_rev_cell': 'P2',
        's2_date_cell': 'N2',
        'data_start': 6,
        'file_prefix': 'PF',
    },
}


def detect_product_line(wb):
    names = wb.sheetnames
    if 'ASUS Commercial PC Info-PF ' in names:
        return 'PF'
    if 'ASUS AiO PC Info-PT' in names:
        return 'AIO(PT)'
    return None

SHEET1_TEMPLATE = [
    ('Air',  'Under 150cm', 'Standard', 150),
    ('',     '',            'Europe',   150),
    ('Sea',  'Under 220cm', 'Standard', 220),
    ('',     '',            'Europe',   220),
    ('',     'Under 210cm', 'Standard', 210),
    ('',     '',            'Europe',   210),
    ('',     'Under 200cm', 'Standard', 200),
    ('',     '',            'Europe',   200),
    ('',     'Under 180cm', 'Standard', 180),
    ('',     '',            'Europe',   180),
    ('',     'Under 150cm', 'Standard', 150),
    ('',     '',            'Europe',   150),
]

C = {
    'bg':'#f0f2f5','white':'#ffffff','surface2':'#f8f9fb',
    'blue':'#3b6cb5','blue_light':'#e8f0fb','blue_pale':'#e8f0fb',
    'green':'#2d8a4e','green_light':'#e6f5ec','green_pale':'#e6f5ec',
    'orange':'#c47a28','orange_light':'#fef5e8','orange_pale':'#fef5e8',
    'red':'#c03838','red_light':'#fef0f0','red_pale':'#fef0f0',
    'blue_info':'#3872b8',
    'border':'#e2e6ec','border_dark':'#cdd3dc','grid':'#e2e6ec',
    'text':'#1a2332','text_dim':'#5a6780','text_label':'#5a6780','text_muted':'#94a0b4',
    'input_bg':'#ffffff','input_bd':'#cdd3dc','input_focus':'#3b6cb5',
    'btn_fg':'#ffffff',
    'btn_blue':'#3b6cb5','btn_hover':'#325ea0',
    'btn_green':'#2d8a4e','btn_green_h':'#247840',
    'btn_orange':'#c47a28','btn_orange_h':'#a86820',
    'prog_bg':'#e2e6ec','prog_fg':'#2d8a4e',
    'edited':'#fff3cd','edited_border':'#f0c040',
    'special_bg':'#fef0dc','special_border':'#f0d4a8',
}


# ======================== 工具函數 ========================

def clean_str(raw):
    if pd.isna(raw):
        return ''
    s = str(raw)
    s = ''.join(' ' if unicodedata.category(c) == 'Zs' else c for c in s)
    return s.strip()


def has_rename_marker(raw_str):
    return '=>' in raw_str or '＝＞' in raw_str


def resolve_location(raw):
    s = clean_str(raw)
    if not s:
        return None
    for sep in ['=>', '＝＞']:
        if sep in s:
            n = s.split(sep)[-1].strip()
            for p in ['更名為', '改名為', '改為', '改']:
                if n.startswith(p):
                    n = n[len(p):]
                    break
            return n.strip() or None
    return s


def calc_layer(lim, pc, bh):
    if lim is None or bh is None or bh <= 0:
        return 1
    return max(1, math.floor((lim - PALLET_HEIGHT[pc]) / bh))


def get_desktop():
    if sys.platform == 'win32':
        desktop = Path(os.environ.get('USERPROFILE', '')) / 'Desktop'
    else:
        desktop = Path.home() / 'Desktop'
    desktop.mkdir(parents=True, exist_ok=True)
    return desktop


def find_product_col(hr, pl):
    t = ['PF'] if pl == 'PF' else ['AIO(PT)', 'PT\n(AIO)']
    for i, v in enumerate(hr):
        if not pd.isna(v) and str(v).strip() in t:
            return i
    return None


def _split_loading_codes(cd):
    if '/' in cd:
        return [c.strip() for c in cd.split('/')]
    return [cd]


# ======================== R 版本讀取 ========================

def read_r_file(r_path, product_line=None):
    wb = openpyxl.load_workbook(r_path, data_only=True)
    try:
        if product_line is None:
            product_line = detect_product_line(wb)
            if product_line is None:
                raise ValueError("無法辨識 R 檔案的產品線，找不到對應的 Sheet")
        cfg = PRODUCT_LINE_CONFIG[product_line]
        ws1 = wb[cfg['sheet1']]
        rev_cell = ws1[cfg['rev_cell']].value or ''
        rev_num = int(''.join(c for c in str(rev_cell) if c.isdigit()) or '0')
        date_cell = ws1[cfg['date_cell']].value or ''
        cc_info = {}
        for row in range(cfg['data_start'], ws1.max_row + 1):
            cc = ws1.cell(row=row, column=19).value
            if cc and str(cc).strip():
                cc_key = str(cc).strip()
                models = str(ws1.cell(row=row, column=2).value or '').strip()
                pcs_std = ws1.cell(row=row, column=11).value
                pcs_eu = ws1.cell(row=row + 1, column=11).value if row + 1 <= ws1.max_row else None
                box_h_mm = ws1.cell(row=row, column=22).value
                cc_info[cc_key] = {
                    'models': models, 'row': row,
                    'pcs_layer_std': pcs_std, 'pcs_layer_eu': pcs_eu,
                    'box_height_mm': box_h_mm,
                }
        ws2 = wb[cfg['sheet2']]
        freight_info = {}
        for row in range(13, ws2.max_row + 1):
            model = ws2.cell(row=row, column=1).value
            ptype = ws2.cell(row=row, column=2).value
            gw = ws2.cell(row=row, column=3).value
            if model and str(ptype).strip() == 'Standard':
                freight_info[str(model).strip()] = {
                    'gw': gw, 'std_row': row, 'eu_row': row + 1,
                }
        return {
            'rev': rev_num, 'date': date_cell, 'product_line': product_line,
            'cc_info': cc_info, 'freight_info': freight_info,
            'last_row_s1': ws1.max_row, 'last_row_s2': ws2.max_row,
        }
    finally:
        wb.close()


# ======================== R 版本生成 ========================

def copy_cell_style(src, dst):
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format


def generate_r_version(r_path, tasks, output_dir=None, product_line=None):
    wb = openpyxl.load_workbook(r_path)
    try:
        if product_line is None:
            product_line = detect_product_line(wb)
            if product_line is None:
                raise ValueError("無法辨識 R 檔案的產品線")
        cfg = PRODUCT_LINE_CONFIG[product_line]
        s1_name = cfg['sheet1']
        ws1 = wb[s1_name]
        ws2 = wb[cfg['sheet2']]
        rev_cell = str(ws1[cfg['rev_cell']].value or 'Rev:0')
        old_rev = int(''.join(c for c in rev_cell if c.isdigit()) or '0')
        new_rev = old_rev + 1
        today = datetime.now().strftime('%Y/%-m/%-d') if sys.platform != 'win32' \
            else datetime.now().strftime('%Y/%#m/%#d')
        ws1[cfg['rev_cell']] = f'Rev:{new_rev}'
        ws1[cfg['date_cell']] = f'Date:{today}'
        ws2[cfg['s2_rev_cell']] = f'Rev:{new_rev}'
        ws2[cfg['s2_date_cell']] = f'Date:{today}'
        s1_next = ws1.max_row + 1
        s2_next = ws2.max_row + 1
        ref_row_s1 = None
        for row in range(6, ws1.max_row + 1):
            if ws1.cell(row=row, column=19).value:
                ref_row_s1 = row
        ref_row_s2 = None
        for row in range(13, ws2.max_row + 1):
            if ws2.cell(row=row, column=1).value and \
               str(ws2.cell(row=row, column=2).value).strip() == 'Standard':
                ref_row_s2 = row
        for task in tasks:
            if task['type'] == 'new':
                _add_new_cc_sheet1(ws1, task, s1_next, ref_row_s1, product_line)
                _add_new_cc_sheet2(ws2, task, s2_next, s1_next, ref_row_s2, s1_name)
                s1_next += 12
                s2_next += 2
            elif task['type'] == 'existing':
                _add_existing_cc(ws1, ws2, task, s2_next, ref_row_s2)
                s2_next += 2
        if output_dir is None:
            output_dir = get_desktop()
        out_name = f'{cfg["file_prefix"]}_Pallet_info_R{new_rev}_0_{datetime.now().strftime("%Y%m%d")}.xlsx'
        out_path = Path(output_dir) / out_name
        wb.save(str(out_path))
        return str(out_path), new_rev
    finally:
        wb.close()


def _add_new_cc_sheet1(ws, task, start_row, ref_row, product_line='AIO(PT)'):
    box_h_mm = task['box_height_mm']
    pcs_std = task['pcs_layer_std']
    pcs_eu = task['pcs_layer_eu']
    pcs_carton = task['pcs_carton']
    h_formula_ph = 14.1
    for i, (air_sea, limit_text, pallet_type, limit_cm) in enumerate(SHEET1_TEMPLATE):
        r = start_row + i
        is_std = (pallet_type == 'Standard')
        pallet_h = 14.1
        w = 100 if is_std else 80
        pcs_layer = pcs_std if is_std else pcs_eu
        layers = max(1, math.floor((limit_cm - pallet_h) / task['box_height_cm']))
        if ref_row:
            ref_offset = ref_row + i
            for col in range(1, 23):
                src = ws.cell(row=ref_offset, column=col)
                dst = ws.cell(row=r, column=col)
                copy_cell_style(src, dst)
        ws.cell(row=r, column=2, value=task['models'])
        ws.cell(row=r, column=3, value=air_sea if air_sea else None)
        ws.cell(row=r, column=4, value=limit_text if limit_text else None)
        ws.cell(row=r, column=5, value=pallet_type)
        ws.cell(row=r, column=6, value=120)
        ws.cell(row=r, column=7, value=w)
        ws.cell(row=r, column=8, value=f'=($V${start_row}*L{r}/10)+{h_formula_ph}')
        ws.cell(row=r, column=9, value=f'=F{r}*G{r}*H{r}/6000')
        ws.cell(row=r, column=10, value=pcs_carton)
        ws.cell(row=r, column=11, value=pcs_layer)
        ws.cell(row=r, column=12, value=layers)
        ws.cell(row=r, column=13, value=f'=K{r}*L{r}')
        ws.cell(row=r, column=14, value=f'=CONCATENATE(K{r},"*",L{r}," (layer)")')
        ws.cell(row=r, column=15, value=f'=(V{start_row}*L{r}/10)+18.5')
        ws.cell(row=r, column=16, value=f'=IF(O{r}<150,L{r},L{r}-1)')
        ws.cell(row=r, column=17, value=f'=M{r}')
        ws.cell(row=r, column=18, value=f'=CONCATENATE(K{r},"*",P{r}," (layer)")')
        if i == 0:
            ws.cell(row=r, column=19, value=task['cc'])
            ws.cell(row=r, column=20, value=task['carton_l_mm'])
            ws.cell(row=r, column=21, value=task['carton_w_mm'])
            ws.cell(row=r, column=22, value=box_h_mm)


def _add_new_cc_sheet2(ws, task, start_row, s1_start, ref_row, s1_name):
    pcs_std = task['pcs_layer_std']
    pcs_eu = task['pcs_layer_eu']
    for i, (ptype, pcs_layer) in enumerate([('Standard', pcs_std), ('Europe', pcs_eu)]):
        r = start_row + i
        if ref_row:
            src_row = ref_row + i
            for col in range(1, 17):
                src = ws.cell(row=src_row, column=col)
                dst = ws.cell(row=r, column=col)
                copy_cell_style(src, dst)
        if i == 0:
            ws.cell(row=r, column=1, value=task['models'])
        else:
            ws.cell(row=r, column=1, value=f'=A{start_row}')
        ws.cell(row=r, column=2, value=ptype)
        ws.cell(row=r, column=3, value=task['gw'])
        ws.cell(row=r, column=4, value=1)
        ws.cell(row=r, column=5, value=pcs_layer)
        s1_row = s1_start + i
        ws.cell(row=r, column=6, value=f"='{s1_name}'!$M${s1_row}")
        sea_rows = {
            220: s1_start + 2 + i, 210: s1_start + 4 + i,
            200: s1_start + 6 + i, 180: s1_start + 8 + i,
            150: s1_start + 10 + i,
        }
        sea_formula = (
            f"=IF($G$7=\"Sea 220cm\",'{s1_name}'!M${sea_rows[220]},"
            f"IF($G$7=\"Sea 210cm\",'{s1_name}'!M${sea_rows[210]},"
            f"IF($G$7=\"Sea 200cm\",'{s1_name}'!M${sea_rows[200]},"
            f"IF($G$7=\"Sea 180cm\",'{s1_name}'!M${sea_rows[180]},"
            f"IF($G$7=\"Sea 150cm\",'{s1_name}'!M${sea_rows[150]},"
            f"\"N/A\")))))"
        )
        ws.cell(row=r, column=7, value=sea_formula)
        ws.cell(row=r, column=8, value=f"='{s1_name}'!I${s1_row}")
        ws.cell(row=r, column=9, value=f'=H{r}')
        ws.cell(row=r, column=10, value=f'=IF(I{r}<K{r},"<",">")')
        ws.cell(row=r, column=11, value=f'=(C{r}*F{r})+30')
        ws.cell(row=r, column=12, value=f'=IF($I{r}>$K{r},($I{r}/$F{r})*$D$3,($K{r}/$F{r})*$D$3)')
        ws.cell(row=r, column=13, value=f'=$D$5/N{r}')
        ws.cell(row=r, column=14, value=f'=G{r}*{"10" if i==0 else "11"}*N$7')
        ws.cell(row=r, column=15, value=f'=$D$7/P{r}')
        ws.cell(row=r, column=16, value=f'=G{r}*{"21" if i==0 else "24"}*P$7')


def _add_existing_cc(ws1, ws2, task, s2_next, ref_row_s2):
    cc = task['cc']
    for row in range(6, ws1.max_row + 1):
        cell_cc = ws1.cell(row=row, column=19).value
        if cell_cc and str(cell_cc).strip() == cc:
            old_models = str(ws1.cell(row=row, column=2).value or '')
            new_models = f"{old_models}/{task['models']}"
            for r in range(row, min(row + 12, ws1.max_row + 1)):
                m_cell = ws1.cell(row=r, column=2)
                if m_cell.value and str(m_cell.value).strip():
                    ws1.cell(row=r, column=2, value=new_models)
            pcs_std = ws1.cell(row=row, column=11).value or 10
            pcs_eu = ws1.cell(row=row + 1, column=11).value or 8
            break
    else:
        return
    existing_gw = task.get('gw')
    if not existing_gw:
        for row in range(13, ws2.max_row + 1):
            m = ws2.cell(row=row, column=1).value
            if m and old_models.split('/')[0] in str(m):
                existing_gw = ws2.cell(row=row, column=3).value
                break
        if not existing_gw:
            existing_gw = 0
    for i, (ptype, pcs_layer) in enumerate([('Standard', pcs_std), ('Europe', pcs_eu)]):
        r = s2_next + i
        if ref_row_s2:
            src_row = ref_row_s2 + i
            for col in range(1, 17):
                src = ws2.cell(row=src_row, column=col)
                dst = ws2.cell(row=r, column=col)
                copy_cell_style(src, dst)
        if i == 0:
            ws2.cell(row=r, column=1, value=task['models'])
        else:
            ws2.cell(row=r, column=1, value=f'=A{s2_next}')
        ws2.cell(row=r, column=2, value=ptype)
        ws2.cell(row=r, column=3, value=existing_gw)
        ws2.cell(row=r, column=4, value=1)
        ws2.cell(row=r, column=5, value=pcs_layer)
        air_layers = max(1, math.floor((150 - 14.1) / task.get('box_height_cm', 50)))
        sea_layers = max(1, math.floor((220 - 14.1) / task.get('box_height_cm', 50)))
        ws2.cell(row=r, column=6, value=pcs_layer * air_layers)
        ws2.cell(row=r, column=7, value=pcs_layer * sea_layers)
        h_air = 14.1 + task.get('box_height_cm', 50) * air_layers
        vw = 120 * (100 if i == 0 else 80) * h_air / 6000
        ws2.cell(row=r, column=8, value=round(vw, 2))
        ws2.cell(row=r, column=9, value=f'=H{r}')
        ws2.cell(row=r, column=10, value=f'=IF(I{r}<K{r},"<",">")')
        ws2.cell(row=r, column=11, value=f'=(C{r}*F{r})+30')
        ws2.cell(row=r, column=12, value=f'=IF($I{r}>$K{r},($I{r}/$F{r})*$D$3,($K{r}/$F{r})*$D$3)')
        ws2.cell(row=r, column=13, value=f'=$D$5/N{r}')
        ws2.cell(row=r, column=14, value=f'=G{r}*{"10" if i==0 else "11"}*N$7')
        ws2.cell(row=r, column=15, value=f'=$D$7/P{r}')
        ws2.cell(row=r, column=16, value=f'=G{r}*{"21" if i==0 else "24"}*P$7')


# ======================== 刪除線偵測 ========================

def get_strikethrough_locations(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=False, read_only=True)
    try:
        strike_set = set()
        for sname in ['歐規棧板', '海空運限高', '陸運', 'EPAL-Amazon', '日字-實木. EPAL#2']:
            if sname not in wb.sheetnames:
                continue
            ws = wb[sname]
            for row in ws.iter_rows(min_row=3, max_col=2):
                if len(row) < 2:
                    continue
                cell = row[1]
                if cell.value and cell.font and cell.font.strikethrough:
                    raw = clean_str(cell.value)
                    if raw:
                        strike_set.add(raw)
        return strike_set
    finally:
        wb.close()


# ======================== CSV 生成核心 ========================

def collect_layer_rules(excel_path, product_line, box_height):
    strike_raw_set = get_strikethrough_locations(excel_path)
    needed_sheets = ['海空運限高', '陸運', '歐規棧板', 'EPAL-Amazon', '日字-實木. EPAL#2']
    try:
        sheets = pd.read_excel(excel_path, sheet_name=needed_sheets, header=None)
    except ValueError:
        sheets = pd.read_excel(excel_path, sheet_name=None, header=None)
    bh = float(box_height) if box_height else None
    anomalies = []

    def check_strike(raw_str):
        return raw_str in strike_raw_set

    def process_location(raw_val, sheet_name, pf_col, row_series, extra_fields=None):
        raw_loc = clean_str(raw_val)
        loc = resolve_location(raw_val)
        if not loc:
            return None, None, True
        pf_val = str(row_series.iloc[pf_col]).strip() if pf_col is not None else ''
        if 'V' not in pf_val:
            return None, None, True
        ef = extra_fields or {}
        if check_strike(raw_loc):
            anomalies.append({
                'location': raw_loc, 'sheet': sheet_name,
                'pallet_type': ef.get('pallet_type', ''),
                'loading_code': ef.get('loading_code', ''),
                'limitation': ef.get('limitation', ''),
                'reason': '有刪除線(已停用)，已排除', 'type': 'strikethrough'})
            return None, None, True
        if has_rename_marker(raw_loc):
            anomalies.append({
                'location': raw_loc, 'sheet': sheet_name,
                'pallet_type': ef.get('pallet_type', ''),
                'loading_code': ef.get('loading_code', ''),
                'limitation': ef.get('limitation', ''),
                'reason': f'含改名標記(=>)，解析為: {loc}', 'type': 'renamed'})
        return loc, raw_loc, False

    sea_air_data = {}
    sa = sheets.get('海空運限高')
    if sa is not None:
        pf = find_product_col(sa.iloc[1].tolist(), product_line)
        if pf is not None:
            for i in range(2, len(sa)):
                r = sa.iloc[i]
                pt = clean_str(r.iloc[9])
                cd = clean_str(r.iloc[11])
                lim_raw = r.iloc[10]
                lim_str = clean_str(lim_raw) if not pd.isna(lim_raw) else ''
                loc, raw_loc, skip = process_location(
                    r.iloc[1], '海空運限高', pf, r,
                    {'pallet_type': pt, 'loading_code': cd, 'limitation': lim_str})
                if skip:
                    continue
                if pt in IGNORED_PALLET_TYPES:
                    anomalies.append({'location': loc, 'sheet': '海空運限高',
                        'pallet_type': pt, 'loading_code': cd, 'limitation': lim_str,
                        'reason': f'Pallet type={pt}，已設定忽略', 'type': 'ignored'})
                    continue
                if pt and pt not in PALLET_TYPE_TO_CODE:
                    anomalies.append({'location': loc, 'sheet': '海空運限高',
                        'pallet_type': pt, 'loading_code': cd, 'limitation': lim_str,
                        'reason': f'未知的 Pallet type="{pt}"', 'type': 'unknown'})
                if cd:
                    for sc in _split_loading_codes(cd):
                        if sc and sc not in LOADING_TYPE_MAP:
                            anomalies.append({'location': loc, 'sheet': '海空運限高',
                                'pallet_type': pt, 'loading_code': cd, 'limitation': lim_str,
                                'reason': f'未知的堆疊方式Code="{sc}"', 'type': 'unknown'})
                try:
                    lv = float(lim_raw)
                except (ValueError, TypeError):
                    lv = None
                    if lim_str:
                        anomalies.append({'location': loc, 'sheet': '海空運限高',
                            'pallet_type': pt, 'loading_code': cd, 'limitation': lim_str,
                            'reason': f'限高無法解析: "{lim_str}"', 'type': 'parse_error'})
                pc = PALLET_TYPE_TO_CODE.get(pt, 'EU')
                sea_air_data.setdefault(loc, []).append((cd, pc, lv))

    land_data, land_lim = {}, {}
    ls = sheets.get('陸運')
    if ls is not None:
        pf = find_product_col(ls.iloc[1].tolist(), product_line)
        if pf is not None:
            for i in range(2, len(ls)):
                r = ls.iloc[i]
                pt = clean_str(r.iloc[8])
                loc, raw_loc, skip = process_location(
                    r.iloc[1], '陸運', pf, r,
                    {'pallet_type': pt, 'loading_code': '陸運'})
                if skip:
                    continue
                if pt in IGNORED_PALLET_TYPES:
                    anomalies.append({'location': loc, 'sheet': '陸運',
                        'pallet_type': pt, 'loading_code': '陸運', 'limitation': '',
                        'reason': f'Pallet type={pt}，已設定忽略', 'type': 'ignored'})
                    continue
                if pt and pt not in PALLET_TYPE_TO_CODE:
                    anomalies.append({'location': loc, 'sheet': '陸運',
                        'pallet_type': pt, 'loading_code': '陸運', 'limitation': '',
                        'reason': f'未知的 Pallet type="{pt}"', 'type': 'unknown'})
                pc = PALLET_TYPE_TO_CODE.get(pt, 'EU')
                land_data.setdefault(loc, []).append(pc)
                try:
                    land_lim[loc] = float(r.iloc[9])
                except (ValueError, TypeError):
                    pass

    eu_locations = set()
    es = sheets.get('歐規棧板')
    if es is not None:
        pf = find_product_col(es.iloc[1].tolist(), product_line)
        if pf is not None:
            for i in range(2, len(es)):
                r = es.iloc[i]
                loc, raw_loc, skip = process_location(
                    r.iloc[1], '歐規棧板', pf, r, {'pallet_type': 'EU'})
                if skip:
                    continue
                eu_locations.add(loc)

    epal_eu_locs, epal_std_locs = set(), set()
    ep = sheets.get('EPAL-Amazon')
    if ep is not None:
        pf = find_product_col(ep.iloc[1].tolist(), product_line)
        if pf is not None:
            for i in range(2, len(ep)):
                r = ep.iloc[i]
                loc, raw_loc, skip = process_location(
                    r.iloc[1], 'EPAL-Amazon', pf, r, {'pallet_type': 'EPAL-EU'})
                if skip:
                    continue
                epal_eu_locs.add(loc)
    ep2 = sheets.get('日字-實木. EPAL#2')
    if ep2 is not None:
        pf = find_product_col(ep2.iloc[1].tolist(), product_line)
        if pf is not None:
            for i in range(2, len(ep2)):
                r = ep2.iloc[i]
                loc, raw_loc, skip = process_location(
                    r.iloc[1], '日字-實木 EPAL#2', pf, r, {'pallet_type': 'EPAL-STD'})
                if skip:
                    continue
                epal_std_locs.add(loc)

    rule_map = {}
    def add_rule(pc, lt, lim, source):
        actual_lim = float(lim) if lim is not None else \
            float(DEFAULT_AIR_LIMIT if lt == 1 else DEFAULT_SEA_LIMIT)
        key = (pc, lt, actual_lim)
        if key not in rule_map:
            rule_map[key] = {
                'pallet_code': pc, 'loading_type': lt,
                'loading_name': LOADING_NAME.get(lt, str(lt)),
                'limitation': actual_lim, 'is_default': (lim is None),
                'calc_layer': calc_layer(actual_lim, pc, bh), 'source': source}

    for loc in eu_locations:
        e = sea_air_data.get(loc, [])
        al = [x[2] for x in e if '空運' in x[0] and x[1] == 'EU' and x[2] is not None]
        sl = [x[2] for x in e if '海運' in x[0] and '散貨' not in x[0]
              and x[1] == 'EU' and x[2] is not None]
        if al:
            for a in al:
                add_rule('EU', 1, a, '歐規棧板')
        else:
            sea_min = min(sl) if sl else None
            if sea_min is not None and sea_min <= DEFAULT_AIR_LIMIT:
                add_rule('EU', 1, sea_min, '海空運限高(空運跟隨海運)')
            else:
                add_rule('EU', 1, None, '歐規棧板')
        for s in sl:
            add_rule('EU', 2, s, '海空運限高')
        if not sl:
            add_rule('EU', 2, None, '歐規棧板(預設)')

    for loc, entries in sea_air_data.items():
        for cd, pc, lv in entries:
            if pc == 'EU' and loc in eu_locations:
                if cd == '海運散貨':
                    add_rule('EU', 6, lv, '海空運限高')
                continue
            for sc in _split_loading_codes(cd):
                lt = LOADING_TYPE_MAP.get(sc)
                if lt is None:
                    continue
                if lt == 4 and loc not in eu_locations:
                    continue
                if pc == 'L7-G' and lt in (2, 4) and lv is not None and lv == 220:
                    continue
                add_rule(pc, lt, lv, '海空運限高')

    for loc, pcs in land_data.items():
        if loc not in eu_locations:
            continue
        for pc in pcs:
            add_rule(pc, 4, land_lim.get(loc), '陸運')

    if epal_eu_locs:
        add_rule('EPAL-EU', 1, None, 'EPAL-Amazon')
        add_rule('EPAL-EU', 2, 175, 'EPAL-Amazon')
    if epal_std_locs:
        add_rule('EPAL-STD', 1, None, '日字EPAL#2')
        add_rule('EPAL-STD', 2, 180, '日字EPAL#2')
    add_rule('EU', 1, None, 'ASTPHQ(固定)')
    add_rule('EU', 2, None, 'ASTPHQ(固定)')

    rules = sorted(rule_map.values(),
                   key=lambda x: (x['pallet_code'], x['loading_type'], x['limitation']))
    loc_data = {
        'sea_air_data': sea_air_data, 'land_data': land_data,
        'land_lim': land_lim, 'eu_locations': eu_locations,
        'epal_eu_locs': epal_eu_locs, 'epal_std_locs': epal_std_locs,
    }
    return rules, loc_data, anomalies


def generate_csv(carton_code, box_height, bl1, bl2, loc_data,
                 layer_lookup, excluded_locations=None):
    hv = float(box_height) if box_height else None
    sad = loc_data['sea_air_data']
    ld, ll = loc_data['land_data'], loc_data['land_lim']
    eu_set = loc_data['eu_locations']
    excl = excluded_locations or set()
    rows = []

    def gl(pc, lt, lim):
        key = (pc, lt, float(lim) if lim is not None else
               float(DEFAULT_AIR_LIMIT if lt == 1 else DEFAULT_SEA_LIMIT))
        if key in layer_lookup:
            return layer_lookup[key]
        if lim is not None and hv and hv > 0:
            return calc_layer(lim, pc, hv)
        return calc_layer(float(DEFAULT_AIR_LIMIT if lt == 1 else DEFAULT_SEA_LIMIT), pc, hv)

    def mr(pc, loc, lt, bl, layer):
        total_height = PALLET_HEIGHT[pc] + hv * layer if hv else None
        return {
            'Carton Code': carton_code, 'Length': PALLET_LENGTH,
            'Width': PALLET_WIDTH[pc], 'Hight': total_height,
            'Ship to locaton': loc, 'Pallet Code': pc,
            'Loading Type': lt, 'Bottom Loading': bl,
            'Bottom Layer': layer, 'Bottom Place Type': 'H',
            'Top Loading': None, 'Top Layer': None, 'Top Place Type': None,
            'Top Cover': None, 'Side Cover': None,
            'Corner Protector Weight': None,
            'MOQ': layer if pc == 'EPAL-EU' else None}

    for loc in sorted(eu_set):
        if loc in excl:
            continue
        e = sad.get(loc, [])
        al = [x[2] for x in e if '空運' in x[0] and x[1] == 'EU' and x[2] is not None]
        sl = [x[2] for x in e if '海運' in x[0] and '散貨' not in x[0]
              and x[1] == 'EU' and x[2] is not None]
        if al:
            air_lim = min(al)
        else:
            sea_min = min(sl) if sl else None
            if sea_min is not None and sea_min <= DEFAULT_AIR_LIMIT:
                air_lim = sea_min
            else:
                air_lim = None
        rows.append(mr('EU', loc, 1, bl1, gl('EU', 1, air_lim)))
        rows.append(mr('EU', loc, 2, bl1, gl('EU', 2, min(sl) if sl else None)))

    for loc in sorted(sad.keys()):
        if loc in excl:
            continue
        for cd, pc, lv in sad[loc]:
            if pc == 'EU' and loc in eu_set:
                if cd == '海運散貨':
                    rows.append(mr('EU', loc, 6, bl1, gl('EU', 6, lv)))
                continue
            bl = bl1 if pc in ('EU', 'EPAL-EU') else bl2
            for sc in _split_loading_codes(cd):
                lt = LOADING_TYPE_MAP.get(sc)
                if lt is None:
                    continue
                if lt == 4 and loc not in eu_set:
                    continue
                if pc == 'L7-G' and lt in (2, 4) and lv is not None and lv == 220:
                    continue
                rows.append(mr(pc, loc, lt, bl, gl(pc, lt, lv)))

    for loc in sorted(ld.keys()):
        if loc in excl or loc not in eu_set:
            continue
        for pc in ld[loc]:
            bl = bl1 if pc in ('EU', 'EPAL-EU') else bl2
            rows.append(mr(pc, loc, 4, bl, gl(pc, 4, ll.get(loc))))

    for loc in sorted(loc_data['epal_eu_locs']):
        if loc in excl:
            continue
        rows.append(mr('EPAL-EU', loc, 1, bl1, gl('EPAL-EU', 1, None)))
        rows.append(mr('EPAL-EU', loc, 2, bl1, gl('EPAL-EU', 2, 175)))

    for loc in sorted(loc_data['epal_std_locs']):
        if loc in excl:
            continue
        rows.append(mr('EPAL-STD', loc, 1, bl2, gl('EPAL-STD', 1, None)))
        rows.append(mr('EPAL-STD', loc, 2, bl2, gl('EPAL-STD', 2, 180)))

    if 'ASTPHQ' not in excl:
        rows.append(mr('EU', 'ASTPHQ', 1, bl1, gl('EU', 1, None)))
        rows.append(mr('EU', 'ASTPHQ', 2, bl1, gl('EU', 2, None)))

    df = pd.DataFrame(rows)
    co = ['Carton Code', 'Length', 'Width', 'Hight', 'Ship to locaton',
          'Pallet Code', 'Loading Type', 'Bottom Loading', 'Bottom Layer',
          'Bottom Place Type', 'Top Loading', 'Top Layer', 'Top Place Type',
          'Top Cover', 'Side Cover', 'Corner Protector Weight', 'MOQ']
    df = df[co].drop_duplicates(
        subset=['Ship to locaton', 'Pallet Code', 'Loading Type'], keep='first')
    df = df.sort_values(
        ['Pallet Code', 'Ship to locaton', 'Loading Type']).reset_index(drop=True)

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'{carton_code}_{ts}.csv'
    op = get_desktop() / filename
    df.to_csv(op, index=False, encoding='big5', errors='replace', lineterminator='\r\n')
    return str(op), len(df)


# ======================== 預覽對話框 (v2.0 完整流程) ========================

class PreviewDialog:
    """新 CC 的特殊棧板 CSV 預覽對話框，複製 v2.0 的完整流程"""

    def __init__(self, parent, cc, rules, loc_data, anomalies, box_height, bl_eu, bl_l7g):
        self.result = None  # 成功時回傳 {layer_lookup, excluded_locs}
        self.rules = rules
        self.loc_data = loc_data
        self.anomalies = anomalies
        self.cc = cc
        self.box_height = box_height
        self.bl_eu = bl_eu
        self.bl_l7g = bl_l7g

        self.anomaly_vars = []

        self.win = tk.Toplevel(parent)
        self.win.title(f"特殊棧板 CSV 預覽 — {cc}")
        self.win.geometry("880x700")
        self.win.minsize(880, 500)
        self.win.configure(bg=C['bg'])
        self.win.transient(parent)
        self.win.grab_set()

        self._build()
        self._show_warnings()

    def _build(self):
        # ── 標題 ──
        top = tk.Frame(self.win, bg=C['white'], padx=12, pady=8)
        top.pack(fill='x')
        tk.Label(top, text=f"📊  {self.cc} — 特殊棧板 CSV 預覽",
                 font=('Microsoft JhengHei UI', 13, 'bold'),
                 bg=C['white'], fg=C['text']).pack(side='left')
        tk.Label(top, text=f"箱高: {self.box_height}cm  BL EU: {self.bl_eu}  BL L7-G: {self.bl_l7g}",
                 font=('Microsoft JhengHei UI', 9),
                 bg=C['white'], fg=C['text_muted']).pack(side='right')
        tk.Frame(self.win, bg=C['border'], height=1).pack(fill='x')

        # ── 主容器 ──
        main = tk.Frame(self.win, bg=C['bg'])
        main.pack(fill='both', expand=True, padx=10, pady=6)

        # ── 異常區 ──
        if self.anomalies:
            af = tk.LabelFrame(main, text=f"  ⚠️  異常偵測 ({len(self.anomalies)} 筆)  ",
                               font=('Microsoft JhengHei UI', 9, 'bold'),
                               bg=C['white'], fg=C['orange'], padx=6, pady=4)
            af.pack(fill='x', pady=(0, 6))

            # 捲動區
            ac = tk.Canvas(af, bg=C['white'], height=min(120, len(self.anomalies) * 22),
                           highlightthickness=0)
            asb = ttk.Scrollbar(af, orient='vertical', command=ac.yview)
            aframe = tk.Frame(ac, bg=C['white'])
            aframe.bind('<Configure>', lambda e: ac.configure(scrollregion=ac.bbox('all')))
            ac.create_window((0, 0), window=aframe, anchor='nw')
            ac.configure(yscrollcommand=asb.set)
            ac.pack(side='left', fill='both', expand=True)
            asb.pack(side='right', fill='y')

            self.anomaly_vars = []
            for i, a in enumerate(self.anomalies):
                atype = a['type']
                # 顏色設定
                if atype in ('strikethrough', 'ignored'):
                    fg, default_checked = C['orange'], True
                elif atype == 'renamed':
                    fg, default_checked = C['blue'], False
                else:  # unknown, parse_error
                    fg, default_checked = C['red'], False

                var = tk.BooleanVar(value=default_checked)
                self.anomaly_vars.append(var)

                row_bg = C['white'] if i % 2 == 0 else C['surface2']
                rf = tk.Frame(aframe, bg=row_bg)
                rf.pack(fill='x')
                tk.Checkbutton(rf, variable=var, bg=row_bg,
                               activebackground=row_bg).pack(side='left')
                tk.Label(rf, text=f"[{a['sheet']}]", font=('Consolas', 8),
                         bg=row_bg, fg=C['text_muted'], width=10, anchor='w').pack(side='left')
                tk.Label(rf, text=a['location'], font=('Consolas', 9, 'bold'),
                         bg=row_bg, fg=fg, width=18, anchor='w').pack(side='left')
                reason_text = a['reason']
                if a.get('pallet_type'):
                    reason_text = f"{a['pallet_type']} | {reason_text}"
                tk.Label(rf, text=reason_text, font=('Microsoft JhengHei UI', 8),
                         bg=row_bg, fg=fg, anchor='w').pack(side='left', fill='x', expand=True)

        # ── 層數預覽表 ──
        lf = tk.LabelFrame(main, text=f"  📊  層數預覽 ({len(self.rules)} 條規則)  — 雙擊層數欄可編輯  ",
                           font=('Microsoft JhengHei UI', 9, 'bold'),
                           bg=C['white'], fg=C['blue'], padx=6, pady=4)
        lf.pack(fill='both', expand=True)

        cols = ('pallet', 'loading', 'limitation', 'layer', 'default', 'source')
        self.tree = ttk.Treeview(lf, columns=cols, show='headings', height=15)
        for c, w, t in [
            ('pallet', 90, '棧板類型'), ('loading', 80, '運輸方式'),
            ('limitation', 80, '限高(cm)'), ('layer', 65, '層數'),
            ('default', 55, '預設'), ('source', 140, '來源'),
        ]:
            self.tree.heading(c, text=t)
            self.tree.column(c, width=w, anchor='center' if c in ('layer', 'default') else 'w')

        tsb = ttk.Scrollbar(lf, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=tsb.set)
        self.tree.pack(side='left', fill='both', expand=True)
        tsb.pack(side='right', fill='y')

        # 填充資料
        self.layer_edits = {}  # (pallet, lt, lim) -> edited_layer
        for r in self.rules:
            iid = self.tree.insert('', 'end', values=(
                r['pallet_code'], r['loading_name'],
                r['limitation'], r['calc_layer'],
                'Y' if r['is_default'] else '',
                r['source'],
            ))
            key = (r['pallet_code'], r['loading_type'], r['limitation'])
            self.layer_edits[key] = r['calc_layer']

            # 預設限高行用橘色標記
            if r['is_default']:
                self.tree.item(iid, tags=('default',))

        self.tree.tag_configure('default', background=C['special_bg'], foreground=C['orange'])
        self.tree.bind('<Double-1>', self._edit_layer)

        # ── 按鈕列 ──
        bf = tk.Frame(self.win, bg=C['bg'], pady=8)
        bf.pack(fill='x')
        tk.Button(bf, text="✅  確認 — 加入待處理清單",
                  font=('Microsoft JhengHei UI', 11, 'bold'),
                  bg=C['btn_green'], fg=C['btn_fg'],
                  activebackground=C['btn_green_h'],
                  relief='flat', padx=24, pady=8, cursor='hand2',
                  command=self._confirm).pack(side='right', padx=14)
        tk.Button(bf, text="取消",
                  font=('Microsoft JhengHei UI', 10),
                  bg=C['border'], fg=C['text_dim'],
                  relief='flat', padx=16, pady=6, cursor='hand2',
                  command=self.win.destroy).pack(side='right')

    def _show_warnings(self):
        """彈出 unknown/parse_error 警告"""
        warnings = [a for a in self.anomalies if a['type'] in ('unknown', 'parse_error')]
        if warnings:
            msg = "⚠️ 發現需要注意的異常：\n\n"
            for w in warnings[:10]:
                msg += f"  [{w['sheet']}] {w['location']}\n    {w['reason']}\n\n"
            if len(warnings) > 10:
                msg += f"  ... 還有 {len(warnings) - 10} 筆\n"
            messagebox.showwarning("異常警告", msg, parent=self.win)

    def _edit_layer(self, event):
        """雙擊編輯層數"""
        item = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if not item or col != '#4':
            return

        vals = self.tree.item(item, 'values')
        bbox = self.tree.bbox(item, 'layer')
        if not bbox:
            return
        x, y, w, h = bbox
        entry = tk.Entry(self.tree, font=('Consolas', 10), justify='center', width=5)
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, vals[3])
        entry.select_range(0, 'end')
        entry.focus_set()

        saved = [False]

        def save(e=None):
            if saved[0]:
                return
            saved[0] = True
            try:
                new_val = int(entry.get())
                if new_val < 1:
                    new_val = 1
            except (ValueError, tk.TclError):
                entry.destroy()
                return
            new_vals = list(vals)
            new_vals[3] = str(new_val)
            self.tree.item(item, values=new_vals)
            pc, lt_name, lim = vals[0], vals[1], float(vals[2])
            lt_map = {v: k for k, v in LOADING_NAME.items()}
            lt = lt_map.get(lt_name, 0)
            key = (pc, lt, lim)
            self.layer_edits[key] = new_val
            tags = list(self.tree.item(item, 'tags'))
            if 'edited' not in tags:
                tags.append('edited')
            self.tree.item(item, tags=tags)
            self.tree.tag_configure('edited', background=C['edited'])
            entry.destroy()

        def cancel(e=None):
            if not saved[0]:
                saved[0] = True
                entry.destroy()

        entry.bind('<Return>', save)
        entry.bind('<FocusOut>', save)
        entry.bind('<Escape>', cancel)

    def _confirm(self):
        """確認並回傳結果"""
        # 收集排除的 locations
        excluded = set()
        for i, a in enumerate(self.anomalies):
            if self.anomaly_vars[i].get():
                loc = resolve_location(a['location'])
                if loc and a['type'] not in ('strikethrough', 'ignored'):
                    excluded.add(loc)

        self.result = {
            'layer_lookup': dict(self.layer_edits),
            'excluded_locs': excluded,
            'rule_count': len(self.rules),
            'anomaly_count': len(self.anomalies),
        }
        self.win.destroy()


# ======================== GUI v3.0 批次模式 ========================

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("棧板包裝資料處理工具 v3.0")
        self.root.geometry("960x900")
        self.root.minsize(960, 700)
        self.root.resizable(True, True)
        self.root.configure(bg=C['bg'])
        self.r_info = {}         # {'AIO(PT)': r_info, 'PF': r_info}
        self.r_path = {}         # {'AIO(PT)': path, 'PF': path}
        self.task_list = []
        self._running = False
        self._cc_is_new = None
        self._cc_product_line = None  # 自動偵測的產品線
        self._build()

    def _section(self, parent, title, ck='blue'):
        color, light = C[ck], C[f'{ck}_light']
        outer = tk.Frame(parent, bg=C['bg'])
        outer.pack(fill='x', padx=14, pady=(0, 6))
        hdr = tk.Frame(outer, bg=light, highlightbackground=C['border'],
                       highlightthickness=1)
        hdr.pack(fill='x')
        tk.Frame(hdr, bg=color, width=4).pack(side='left', fill='y')
        tk.Label(hdr, text=f"  {title}", font=('Microsoft JhengHei UI', 10, 'bold'),
                 bg=light, fg=color, pady=5).pack(side='left')
        bw = tk.Frame(outer, bg=C['border'])
        bw.pack(fill='both', expand=True)
        body = tk.Frame(bw, bg=C['white'])
        body.pack(fill='both', expand=True, padx=1, pady=(0, 1))
        content = tk.Frame(body, bg=C['white'], padx=10, pady=5)
        content.pack(fill='both', expand=True)
        return content

    def _field_row(self, parent, label, default='', combo=None, idx=0, width=28):
        bg = C['white'] if idx % 2 == 0 else C['surface2']
        row = tk.Frame(parent, bg=bg)
        row.pack(fill='x')
        lf = tk.Frame(row, bg=C['surface2'], width=260)
        lf.pack(side='left', fill='y')
        lf.pack_propagate(False)
        tk.Label(lf, text=label, font=('Microsoft JhengHei UI', 9),
                 bg=C['surface2'], fg=C['text_dim'],
                 anchor='e', padx=12, pady=5).pack(fill='both', expand=True)
        tk.Frame(row, bg=C['grid'], width=1).pack(side='left', fill='y')
        vf = tk.Frame(row, bg=bg, padx=8, pady=2)
        vf.pack(side='left', fill='both', expand=True)
        var = tk.StringVar(value=default)
        if combo:
            ttk.Combobox(vf, textvariable=var, values=combo,
                         state='readonly', width=width, font=('Consolas', 10)).pack(side='left', pady=2)
        else:
            tk.Entry(vf, textvariable=var, width=width, font=('Consolas', 10),
                     bg=C['input_bg'], fg=C['text'], relief='solid', bd=1,
                     highlightthickness=1, highlightcolor=C['input_focus'],
                     highlightbackground=C['input_bd']).pack(side='left', pady=2)
        return var

    def _group_label(self, parent, text):
        gf = tk.Frame(parent, bg=C['surface2'])
        gf.pack(fill='x')
        tk.Label(gf, text=f"  ── {text} ──", font=('Microsoft JhengHei UI', 8),
                 bg=C['surface2'], fg=C['text_muted'], anchor='w',
                 pady=3, padx=4).pack(fill='x')
        tk.Frame(parent, bg=C['border'], height=1).pack(fill='x')

    def _ref_row(self, parent, pairs, idx=0):
        bg = C['white'] if idx % 2 == 0 else C['surface2']
        row = tk.Frame(parent, bg=bg)
        row.pack(fill='x')
        for i, (k, v) in enumerate(pairs):
            if i > 0:
                tk.Frame(row, bg=C['grid'], width=1).pack(side='left', fill='y')
            cell = tk.Frame(row, bg=bg, padx=6, pady=2)
            cell.pack(side='left', fill='both', expand=True)
            if k:
                tk.Label(cell, text=k, font=('Microsoft JhengHei UI', 7),
                         bg=bg, fg=C['text_muted']).pack(anchor='w')
            tk.Label(cell, text=v, font=('Microsoft JhengHei UI', 8, 'bold'),
                     bg=bg, fg=C['text']).pack(anchor='w')
        tk.Frame(parent, bg=C['grid'], height=1).pack(fill='x')

    def _pill(self, parent, text, ck):
        color, light = C[ck], C[f'{ck}_light']
        f = tk.Frame(parent, bg=light, highlightbackground=C['border'],
                     highlightthickness=1, padx=6, pady=1)
        f.pack(side='left', padx=2)
        tk.Label(f, text=text, font=('Microsoft JhengHei UI', 7, 'bold'),
                 bg=light, fg=color).pack()

    def _bind_mousewheel(self, canvas):
        def _on_wheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        def _on_linux_up(event):
            canvas.yview_scroll(-1, "units")
        def _on_linux_down(event):
            canvas.yview_scroll(1, "units")
        def _bind(event):
            canvas.bind_all("<MouseWheel>", _on_wheel)
            canvas.bind_all("<Button-4>", _on_linux_up)
            canvas.bind_all("<Button-5>", _on_linux_down)
        def _unbind(event):
            canvas.unbind_all("<MouseWheel>")
            canvas.unbind_all("<Button-4>")
            canvas.unbind_all("<Button-5>")
        canvas.bind("<Enter>", _bind)
        canvas.bind("<Leave>", _unbind)

    def _build(self):
        # 標題列
        top = tk.Frame(self.root, bg=C['white'], padx=12, pady=8)
        top.pack(fill='x')
        icon_frame = tk.Frame(top, bg=C['blue'], width=38, height=38)
        icon_frame.pack(side='left', padx=(0, 10))
        icon_frame.pack_propagate(False)
        tk.Label(icon_frame, text="📦", font=('Microsoft JhengHei UI', 16),
                 bg=C['blue'], fg=C['btn_fg']).pack(expand=True)
        tf = tk.Frame(top, bg=C['white'])
        tf.pack(side='left')
        tk.Label(tf, text="棧板包裝資料處理工具",
                 font=('Microsoft JhengHei UI', 14, 'bold'),
                 bg=C['white'], fg=C['text']).pack(anchor='w')
        tk.Label(tf, text="Pallet Packing Configuration Tool v3.0 — 批次模式",
                 font=('Microsoft JhengHei UI', 8), bg=C['white'], fg=C['text_muted']).pack(anchor='w')
        flow = tk.Frame(top, bg=C['white'])
        flow.pack(side='right')
        for txt, ck in [("📄Pallet info", 'blue'), ("→", ""), ("➕新增CC", 'green'),
                        ("→", ""), ("🔍預覽", 'orange'), ("→", ""), ("📊產出", 'green')]:
            if txt == "→":
                tk.Label(flow, text="→", font=('Microsoft JhengHei UI', 8),
                         bg=C['white'], fg=C['text_muted']).pack(side='left', padx=2)
            else:
                self._pill(flow, txt, ck)
        tk.Frame(self.root, bg=C['border'], height=1).pack(fill='x')

        # 捲動容器
        container = tk.Frame(self.root, bg=C['bg'])
        container.pack(fill='both', expand=True)
        self.canvas = tk.Canvas(container, bg=C['bg'], highlightthickness=0)
        vsb = ttk.Scrollbar(container, orient='vertical', command=self.canvas.yview)
        self.main = tk.Frame(self.canvas, bg=C['bg'])
        self.main.bind('<Configure>',
                       lambda e: self.canvas.configure(scrollregion=self.canvas.bbox('all')))
        self.canvas.create_window((0, 0), window=self.main, anchor='nw', tags='main_win')
        self.canvas.configure(yscrollcommand=vsb.set)
        def on_resize(e):
            self.canvas.itemconfig('main_win', width=e.width)
        self.canvas.bind('<Configure>', on_resize)
        self._bind_mousewheel(self.canvas)
        self.canvas.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')

        # ── Step 0: Pallet info 檔案選擇（PT + PF）──
        fc = self._section(self.main, "📁  Step 0：選擇 Pallet info 檔案", 'blue')

        # PT Pallet info
        pt_row = tk.Frame(fc, bg=C['white'])
        pt_row.pack(fill='x', pady=2)
        tk.Label(pt_row, text="PT Pallet info", font=('Microsoft JhengHei UI', 9, 'bold'),
                 bg=C['white'], fg=C['text_label'], width=14, anchor='e').pack(side='left')
        self.r_file_pt_var = tk.StringVar()
        tk.Entry(pt_row, textvariable=self.r_file_pt_var, width=44,
                 font=('Consolas', 9), bg=C['input_bg'], fg=C['text'],
                 relief='solid', bd=1, highlightthickness=1,
                 highlightcolor=C['input_focus'],
                 highlightbackground=C['input_bd']).pack(side='left', padx=6)
        tk.Button(pt_row, text="📂 瀏覽", font=('Microsoft JhengHei UI', 9),
                  bg=C['btn_blue'], fg=C['btn_fg'], activebackground=C['btn_hover'],
                  relief='flat', padx=10, pady=2, cursor='hand2',
                  command=lambda: self._browse_r('AIO(PT)')).pack(side='left')
        self.r_info_pt_label = tk.Label(fc, text="", font=('Microsoft JhengHei UI', 8),
                                        bg=C['white'], fg=C['text_dim'])
        self.r_info_pt_label.pack(anchor='w', padx=(0, 0))

        tk.Frame(fc, bg=C['grid'], height=1).pack(fill='x', pady=2)

        # PF Pallet info
        pf_row = tk.Frame(fc, bg=C['white'])
        pf_row.pack(fill='x', pady=2)
        tk.Label(pf_row, text="PF Pallet info", font=('Microsoft JhengHei UI', 9, 'bold'),
                 bg=C['white'], fg=C['text_label'], width=14, anchor='e').pack(side='left')
        self.r_file_pf_var = tk.StringVar()
        tk.Entry(pf_row, textvariable=self.r_file_pf_var, width=44,
                 font=('Consolas', 9), bg=C['input_bg'], fg=C['text'],
                 relief='solid', bd=1, highlightthickness=1,
                 highlightcolor=C['input_focus'],
                 highlightbackground=C['input_bd']).pack(side='left', padx=6)
        tk.Button(pf_row, text="📂 瀏覽", font=('Microsoft JhengHei UI', 9),
                  bg=C['btn_blue'], fg=C['btn_fg'], activebackground=C['btn_hover'],
                  relief='flat', padx=10, pady=2, cursor='hand2',
                  command=lambda: self._browse_r('PF')).pack(side='left')
        self.r_info_pf_label = tk.Label(fc, text="", font=('Microsoft JhengHei UI', 8),
                                        bg=C['white'], fg=C['text_dim'])
        self.r_info_pf_label.pack(anchor='w')

        # ── Step 1: 新增 CC ──
        ic = self._section(self.main, "➕  Step 1：新增 Carton Code", 'green')
        self._group_label(ic, "產品 / 機種")
        self.cc_var = self._field_row(ic, "Carton Code", '', idx=0)
        self.models_var = self._field_row(ic, "Models 機種名稱", '', idx=1)

        # 檢查按鈕 + 狀態
        chk_frame = tk.Frame(ic, bg=C['white'])
        chk_frame.pack(fill='x', pady=(4, 2))
        self.check_btn = tk.Button(
            chk_frame, text="🔍  檢查 CC 新舊",
            font=('Microsoft JhengHei UI', 9, 'bold'), bg=C['btn_orange'], fg=C['btn_fg'],
            activebackground=C['btn_orange_h'], relief='flat', padx=14, pady=4,
            cursor='hand2', command=self._check_cc)
        self.check_btn.pack(side='left', padx=(0, 10))
        self.cc_status_var = tk.StringVar(value="")
        self.cc_status_label = tk.Label(chk_frame, textvariable=self.cc_status_var,
            font=('Microsoft JhengHei UI', 9, 'bold'), bg=C['white'], fg=C['text_dim'])
        self.cc_status_label.pack(side='left')

        # ── 新 CC 專用欄位（預設隱藏）──
        self.new_cc_frame = tk.Frame(ic, bg=C['white'])
        # 不 pack — 預設隱藏，等 _check_cc 判斷後才顯示

        self._group_label(self.new_cc_frame, "箱子規格（新 CC 專用）")
        self.height_var = self._field_row(self.new_cc_frame, "箱高 (cm)", '', idx=0)
        self.carton_l_var = self._field_row(self.new_cc_frame, "箱子尺寸 L (cm)", '', idx=1)
        self.carton_w_var = self._field_row(self.new_cc_frame, "箱子尺寸 W (cm)", '', idx=0)
        self.gw_var = self._field_row(self.new_cc_frame, "G.W. 毛重 (kg)", '', idx=1)
        self.pcs_carton_var = self._field_row(self.new_cc_frame, "Pcs/@Carton 每箱件數", '1', idx=0)

        self._group_label(self.new_cc_frame, "棧板配置")
        self.bl_eu_var = self._field_row(self.new_cc_frame, "Bottom Loading EU / EPAL-EU", '', idx=0)
        self.bl_l7g_var = self._field_row(self.new_cc_frame, "Bottom Loading L7-G / EPAL-STD", '', idx=1)

        self._group_label(self.new_cc_frame, "Flora 特殊棧板檔案（產出特殊棧板 CSV）")
        arow = tk.Frame(self.new_cc_frame, bg=C['white'])
        arow.pack(fill='x', pady=2)
        tk.Label(arow, text="Flora 檔案", font=('Microsoft JhengHei UI', 9),
                 bg=C['white'], fg=C['text_label']).pack(side='left')
        self.a_file_var = tk.StringVar()
        tk.Entry(arow, textvariable=self.a_file_var, width=40,
                 font=('Consolas', 9), bg=C['input_bg'], fg=C['text'],
                 relief='solid', bd=1, highlightthickness=1,
                 highlightcolor=C['input_focus'],
                 highlightbackground=C['input_bd']).pack(side='left', padx=6)
        tk.Button(arow, text="📂 瀏覽", font=('Microsoft JhengHei UI', 9),
                  bg=C['btn_blue'], fg=C['btn_fg'],
                  activebackground=C['btn_hover'],
                  relief='flat', padx=10, pady=2, cursor='hand2',
                  command=self._browse_a).pack(side='left')
        self.a_product_var = self._field_row(self.new_cc_frame, "Flora 檔案產品線", 'PF',
                                             PRODUCT_LINE_OPTIONS, idx=1)

        # 加入清單按鈕
        bf = tk.Frame(ic, bg=C['white'])
        bf.pack(fill='x', pady=(8, 4))
        self.add_btn = tk.Button(
            bf, text="✚  加入待處理清單",
            font=('Microsoft JhengHei UI', 10, 'bold'), bg=C['btn_green'], fg=C['btn_fg'],
            activebackground=C['btn_green_h'], relief='flat', padx=20, pady=6,
            cursor='hand2', command=self._add_to_list)
        self.add_btn.pack()

        # ── 待處理清單 ──
        self.list_section = self._section(self.main, "📋  待處理清單", 'orange')
        self.list_frame = tk.Frame(self.list_section, bg=C['white'])
        self.list_frame.pack(fill='x')
        self.list_empty_label = tk.Label(self.list_frame,
            text="尚未加入任何項目，請在上方填寫後按「加入待處理清單」",
            font=('Microsoft JhengHei UI', 9), bg=C['white'], fg=C['text_muted'], pady=10)
        self.list_empty_label.pack()

        # ── 預覽區 (動態) ──
        self.preview_frame = tk.Frame(self.main, bg=C['bg'])
        self.preview_frame.pack(fill='x')

        # ── 產出按鈕 ──
        bf2 = tk.Frame(self.main, bg=C['bg'])
        bf2.pack(pady=(6, 4))
        self.gen_btn = tk.Button(
            bf2, text="🚀  確認產出",
            font=('Microsoft JhengHei UI', 12, 'bold'), bg=C['btn_blue'], fg=C['btn_fg'],
            activebackground=C['btn_hover'], relief='flat', padx=28, pady=10,
            cursor='hand2', command=self._generate_all, state='disabled')
        self.gen_btn.pack()

        # ── 固定參數 ──
        rc = self._section(self.main, "📋  固定參數 Reference", 'orange')
        self._ref_row(rc, [('運輸類型', '空運:1  海運:2  陸運:4  散貨:6')], 0)
        self._ref_row(rc, [('棧板高(cm)',
                            'EU/L7-G:14.1  EPAL-EU:16.5  EPAL-STD:18.5')], 1)
        self._ref_row(rc, [('預設限高', '空運:150cm  海運:220cm'),
                           ('公式', 'floor((限高−棧板高)÷箱高)')], 2)
        self._ref_row(rc, [('固定加入', 'ASTPHQ → EU 空運+海運'),
                           ('EPAL-EU', 'MOQ = Bottom Layer')], 3)

        # 狀態列
        sf = tk.Frame(self.root, bg=C['surface2'], height=28)
        sf.pack(fill='x', side='bottom')
        self.status_var = tk.StringVar(value="就緒")
        tk.Label(sf, textvariable=self.status_var, font=('Microsoft JhengHei UI', 8),
                 bg=C['surface2'], fg=C['text_dim'], padx=10).pack(side='left')

    # ── 事件處理 ──

    def _detect_cc_product_line(self, cc):
        """從 CC 前綴判斷產品線: PT->AIO(PT), DT->PF"""
        cc_upper = cc.upper()
        if cc_upper.startswith('PT'):
            return 'AIO(PT)'
        elif cc_upper.startswith('DT'):
            return 'PF'
        return None

    def _browse_r(self, product_line):
        fp = filedialog.askopenfilename(
            title=f"選擇 {product_line} Pallet info 檔案",
            filetypes=[("Excel", "*.xlsx")])
        if not fp:
            return
        # 更新對應的路徑和標籤
        if product_line == 'AIO(PT)':
            file_var = self.r_file_pt_var
            info_label = self.r_info_pt_label
        else:
            file_var = self.r_file_pf_var
            info_label = self.r_info_pf_label
        file_var.set(fp)
        try:
            r_info = read_r_file(fp, product_line)
            self.r_info[product_line] = r_info
            self.r_path[product_line] = fp
            rev = r_info['rev']
            cc_count = len(r_info['cc_info'])
            info_label.config(
                text=f"  ✅ Rev: {rev}，{cc_count} 個 CC",
                fg=C['green'])
            self.status_var.set(f"已載入 {product_line} R{rev}")
        except Exception as e:
            self.r_info.pop(product_line, None)
            self.r_path.pop(product_line, None)
            info_label.config(text=f"  ❌ 載入失敗: {e}", fg=C['red'])

    def _browse_a(self):
        fp = filedialog.askopenfilename(
            title="選擇 Flora 特殊棧板檔案",
            filetypes=[("Excel", "*.xlsx")])
        if fp:
            self.a_file_var.set(fp)

    def _check_cc(self):
        """檢查 CC 是新的還是舊的，展開或隱藏對應欄位"""
        cc = self.cc_var.get().strip()
        if not cc:
            messagebox.showerror("錯誤", "請先輸入 Carton Code！")
            return

        # 從 CC 前綴判斷產品線
        pl = self._detect_cc_product_line(cc)
        if pl is None:
            messagebox.showerror("錯誤",
                f"無法從 CC「{cc}」判斷產品線！\n\n"
                f"CC 應以 PT（AIO）或 DT（PF）開頭。")
            return

        # 檢查對應產品線的 Pallet info 是否已載入
        if pl not in self.r_info:
            pl_name = "PT" if pl == 'AIO(PT)' else "PF"
            messagebox.showerror("錯誤",
                f"請先載入 {pl_name} Pallet info 檔案！\n\n"
                f"CC「{cc}」屬於 {pl} 產品線。")
            return

        # 檢查是否已在待處理清單
        for t in self.task_list:
            if t['cc'] == cc:
                messagebox.showerror("錯誤", f"{cc} 已在待處理清單中！")
                return

        r_info = self.r_info[pl]
        is_new = cc not in r_info['cc_info']
        self._cc_is_new = is_new
        self._cc_product_line = pl

        if is_new:
            self.new_cc_frame.pack(fill='x', before=self.add_btn.master)
            self.a_product_var.set(pl)
            self.cc_status_var.set(
                f"🆕 {cc} 是新的 Carton Code（{pl}）\n"
                f"     → 請填寫下方欄位")
            self.cc_status_label.config(fg=C['orange'])
            self.status_var.set(f"{cc} 是新 CC（{pl}），請填寫箱子規格")
        else:
            self.new_cc_frame.pack_forget()
            existing = r_info['cc_info'][cc]
            self.cc_status_var.set(
                f"📎 {cc} 已存在（{pl}）— 現有機種: {existing['models']}\n"
                f"     → 追加新機種（不產 CSV）")
            self.cc_status_label.config(fg=C['blue'])
            self.status_var.set(f"{cc} 是舊 CC（{pl}），只需輸入 Models")

    def _add_to_list(self):
        """驗證並加入待處理清單"""
        cc = self.cc_var.get().strip()
        models = self.models_var.get().strip()
        if not cc:
            messagebox.showerror("錯誤", "請輸入 Carton Code！")
            return
        if not models:
            messagebox.showerror("錯誤", "請輸入 Models 機種名稱！")
            return
        if self._cc_is_new is None:
            messagebox.showerror("錯誤", "請先按「檢查 CC 新舊」確認 CC 類型！")
            return

        pl = self._cc_product_line
        if pl not in self.r_info:
            messagebox.showerror("錯誤", f"請先載入 {pl} 的 Pallet info 檔案！")
            return

        r_info = self.r_info[pl]
        current_is_new = cc not in r_info['cc_info']
        if current_is_new != self._cc_is_new:
            self._check_cc()
            return
        for t in self.task_list:
            if t['cc'] == cc:
                messagebox.showerror("錯誤", f"{cc} 已在待處理清單中！")
                return
        is_new = self._cc_is_new
        if is_new:
            errs = []
            try:
                bh = float(self.height_var.get().strip())
                if bh <= 0: raise ValueError
            except (ValueError, TypeError):
                errs.append("Box Height (cm)")
            try:
                cl = float(self.carton_l_var.get().strip())
            except (ValueError, TypeError):
                errs.append("Carton L (cm)")
            try:
                cw = float(self.carton_w_var.get().strip())
            except (ValueError, TypeError):
                errs.append("Carton W (cm)")
            try:
                gw = float(self.gw_var.get().strip())
            except (ValueError, TypeError):
                errs.append("G.W. (kg)")
            try:
                pcs = int(self.pcs_carton_var.get().strip())
            except (ValueError, TypeError):
                errs.append("Pcs/@Carton")
            try:
                b1 = int(self.bl_eu_var.get().strip())
            except (ValueError, TypeError):
                errs.append("Bottom Loading EU")
            try:
                b2 = int(self.bl_l7g_var.get().strip())
            except (ValueError, TypeError):
                errs.append("Bottom Loading L7-G")
            if errs:
                messagebox.showerror("欄位不完整",
                                     "以下欄位缺少或格式錯誤：\n\n" +
                                     "\n".join(f"  • {e}" for e in errs))
                return
            if b1 >= b2:
                messagebox.showerror("錯誤",
                    f"Bottom Loading 設定錯誤！\n\n"
                    f"EU/EPAL-EU ({b1}) 必須小於 L7-G/EPAL-STD ({b2})")
                return
            if not self.a_file_var.get().strip():
                messagebox.showerror("錯誤", "新 CC 需要 Flora 特殊棧板檔案才能產出 CSV！")
                return

            a_file = self.a_file_var.get().strip()
            a_pl = self.a_product_var.get()

            # ── 執行預覽流程 ──
            self.status_var.set(f"正在分析 {cc}...")
            self.root.update_idletasks()

            try:
                rules, loc_data, anomalies = collect_layer_rules(a_file, a_pl, str(bh))
            except Exception as e:
                messagebox.showerror("分析失敗", f"Flora 檔案分析失敗：\n\n{e}")
                return

            # 開啟預覽對話框
            dlg = PreviewDialog(self.root, cc, rules, loc_data, anomalies, bh, b1, b2)
            self.root.wait_window(dlg.win)

            if dlg.result is None:
                self.status_var.set(f"{cc} 預覽已取消")
                return

            task = {
                'type': 'new', 'cc': cc, 'models': models,
                'product_line': pl,
                'box_height_cm': bh, 'box_height_mm': int(bh * 10),
                'carton_l_mm': int(cl * 10), 'carton_w_mm': int(cw * 10),
                'pcs_carton': pcs, 'pcs_layer_std': b2, 'pcs_layer_eu': b1,
                'gw': gw, 'bl_eu': b1, 'bl_l7g': b2,
                'a_file': a_file,
                'a_product_line': a_pl,
                'preview_layer_lookup': dlg.result['layer_lookup'],
                'preview_excluded_locs': dlg.result['excluded_locs'],
                'preview_loc_data': loc_data,
            }
        else:
            existing = r_info['cc_info'][cc]
            box_h_mm = existing.get('box_height_mm')
            box_h_cm = box_h_mm / 10.0 if box_h_mm else 50.0
            task = {
                'type': 'existing', 'cc': cc, 'models': models,
                'product_line': pl,
                'existing_models': existing['models'],
                'box_height_cm': box_h_cm,
            }
        self.task_list.append(task)
        self._refresh_list()
        self.cc_var.set('')
        self.models_var.set('')
        self._cc_is_new = None
        self.cc_status_var.set("")
        self.new_cc_frame.pack_forget()
        if is_new:
            self.height_var.set('')
            self.carton_l_var.set('')
            self.carton_w_var.set('')
            self.gw_var.set('')
            self.bl_eu_var.set('')
            self.bl_l7g_var.set('')
        self.status_var.set(
            f"已加入 {cc}（{'新' if is_new else '舊'} CC），待處理共 {len(self.task_list)} 筆")

    def _refresh_list(self):
        for w in self.list_frame.winfo_children():
            w.destroy()
        if not self.task_list:
            tk.Label(self.list_frame,
                text="尚未加入任何項目",
                font=('Microsoft JhengHei UI', 9), bg=C['white'], fg=C['text_muted'],
                pady=10).pack()
            self.gen_btn.config(state='disabled', bg=C['border'])
            return
        hdr = tk.Frame(self.list_frame, bg=C['surface2'])
        hdr.pack(fill='x')
        tk.Frame(self.list_frame, bg=C['border'], height=2).pack(fill='x')
        for txt, w in [("#", 3), ("CC", 10), ("產品線", 6), ("類型", 6), ("Models", 22), ("刪除", 5)]:
            tk.Label(hdr, text=txt, font=('Microsoft JhengHei UI', 8, 'bold'),
                     bg=C['surface2'], fg=C['text_muted'], width=w, pady=4,
                     anchor='w', padx=4).pack(side='left')
        for idx, task in enumerate(self.task_list):
            bg = C['white'] if idx % 2 == 0 else C['surface2']
            is_new = task['type'] == 'new'
            pl = task.get('product_line', '')
            pl_short = 'PT' if pl == 'AIO(PT)' else 'PF'
            row = tk.Frame(self.list_frame, bg=bg)
            row.pack(fill='x')
            tk.Label(row, text=str(idx + 1), font=('Consolas', 9),
                     bg=bg, fg=C['text_dim'], width=3, anchor='w', padx=4).pack(side='left')
            tk.Label(row, text=task['cc'], font=('Consolas', 10, 'bold'),
                     bg=bg, fg=C['text'], width=10, anchor='w').pack(side='left')
            tk.Label(row, text=pl_short, font=('Consolas', 9),
                     bg=bg, fg=C['text_dim'], width=6, anchor='w').pack(side='left')
            type_bg = C['orange_light'] if is_new else C['blue_light']
            type_fg = C['orange'] if is_new else C['blue']
            type_text = "🆕 新" if is_new else "📎 舊"
            tframe = tk.Frame(row, bg=type_bg, padx=4, pady=1)
            tframe.pack(side='left', padx=4)
            tk.Label(tframe, text=type_text, font=('Microsoft JhengHei UI', 8, 'bold'),
                     bg=type_bg, fg=type_fg).pack()
            model_text = task['models']
            if not is_new:
                model_text = f"{task['existing_models']} + {task['models']}"
            tk.Label(row, text=model_text, font=('Consolas', 9),
                     bg=bg, fg=C['text'], width=22, anchor='w').pack(side='left')
            def make_del(i):
                return lambda: self._remove_task(i)
            tk.Button(row, text="✕", font=('Microsoft JhengHei UI', 8),
                      bg=bg, fg=C['red'], relief='flat', cursor='hand2',
                      command=make_del(idx), padx=4).pack(side='left')

        new_count = sum(1 for t in self.task_list if t['type'] == 'new')
        old_count = sum(1 for t in self.task_list if t['type'] == 'existing')
        parts = []
        if new_count:
            parts.append(f"{new_count} 個新 CC")
        if old_count:
            parts.append(f"{old_count} 個舊 CC")
        btn_text = f"🚀  確認產出 — {' + '.join(parts)}"
        self.gen_btn.config(state='normal', bg=C['btn_blue'], text=btn_text)

    def _remove_task(self, idx):
        if 0 <= idx < len(self.task_list):
            removed = self.task_list.pop(idx)
            self._refresh_list()
            self.status_var.set(f"已移除 {removed['cc']}")

    def _generate_all(self):
        if not self.task_list:
            return
        if self._running:
            return
        self._running = True
        self.gen_btn.config(state='disabled', bg=C['border'])
        self.status_var.set("產出中...")

        tasks_snapshot = list(self.task_list)
        r_path_snapshot = dict(self.r_path)

        def work():
            try:
                results = []

                # 按產品線分組
                tasks_by_pl = {}
                for task in tasks_snapshot:
                    pl = task.get('product_line', 'AIO(PT)')
                    tasks_by_pl.setdefault(pl, []).append(task)

                # 各產品線分別生成新版 Pallet info
                new_r_paths = {}
                for pl, pl_tasks in tasks_by_pl.items():
                    r_path = r_path_snapshot.get(pl)
                    if not r_path:
                        results.append(f"⚠️ {pl} — Pallet info 未載入，跳過")
                        continue
                    try:
                        out_r, new_rev = generate_r_version(
                            r_path, pl_tasks, product_line=pl)
                        pl_short = 'PT' if pl == 'AIO(PT)' else 'PF'
                        results.append(f"✅ {pl_short} R{new_rev} → {os.path.basename(out_r)}")
                        new_r_paths[pl] = (out_r, new_rev)
                    except Exception as e:
                        results.append(f"❌ {pl} Pallet info 更新失敗: {e}")

                # 新 CC 產出特殊棧板 CSV
                for task in tasks_snapshot:
                    if task['type'] != 'new':
                        continue
                    cc = task['cc']
                    bh = str(task['box_height_cm'])
                    bl1, bl2 = task['bl_eu'], task['bl_l7g']

                    layer_lookup = task.get('preview_layer_lookup', {})
                    excluded_locs = task.get('preview_excluded_locs', set())
                    loc_data = task.get('preview_loc_data')

                    if not loc_data:
                        results.append(f"⚠️ {cc} — 缺少預覽資料，跳過 CSV")
                        continue

                    try:
                        csv_path, csv_cnt = generate_csv(
                            cc, bh, bl1, bl2, loc_data, layer_lookup, excluded_locs)
                        results.append(
                            f"✅ {cc} CSV → {os.path.basename(csv_path)}（{csv_cnt} 筆）")
                    except Exception as e:
                        results.append(f"❌ {cc} CSV 失敗: {e}")

                msg = "產出完成！\n\n" + "\n".join(results) + "\n\n檔案已儲存到桌面"
                self.root.after(0, lambda: messagebox.showinfo("✅ 產出完成", msg))

                new_cc_count = sum(1 for t in tasks_snapshot if t['type'] == 'new')
                old_cc_count = sum(1 for t in tasks_snapshot if t['type'] == 'existing')
                self.root.after(0, lambda: self.status_var.set(
                    f"✅ 完成 — {new_cc_count} 個新 CC + {old_cc_count} 個舊 CC"))

                def _post_generate():
                    self.task_list.clear()
                    self._refresh_list()
                    for pl, (out_r, new_rev) in new_r_paths.items():
                        self.r_path[pl] = out_r
                        if pl == 'AIO(PT)':
                            self.r_file_pt_var.set(out_r)
                        else:
                            self.r_file_pf_var.set(out_r)
                        try:
                            self.r_info[pl] = read_r_file(out_r, pl)
                            cc_count = len(self.r_info[pl]['cc_info'])
                            label = self.r_info_pt_label if pl == 'AIO(PT)' else self.r_info_pf_label
                            label.config(
                                text=f"  ✅ Rev: {new_rev}，{cc_count} 個 CC",
                                fg=C['green'])
                        except Exception:
                            pass
                    self._running = False
                    self.gen_btn.config(
                        state='normal' if self.task_list else 'disabled',
                        bg=C['btn_blue'] if self.task_list else C['border'])

                self.root.after(0, _post_generate)

            except Exception as e:
                tb = traceback.format_exc()
                def _on_error():
                    messagebox.showerror("產出失敗", f"{e}\n\n{tb}")
                    self.status_var.set("❌ 產出失敗")
                    self._running = False
                    self.gen_btn.config(
                        state='normal' if self.task_list else 'disabled',
                        bg=C['btn_blue'] if self.task_list else C['border'])
                self.root.after(0, _on_error)

        threading.Thread(target=work, daemon=True).start()


def main():
    root = tk.Tk()

    import tkinter.font as tkfont
    default_font = tkfont.nametofont("TkDefaultFont")
    default_font.configure(family='Microsoft JhengHei UI', size=9)
    root.option_add("*Font", default_font)

    style = ttk.Style()
    for theme in ['clam', 'vista', 'winnative']:
        if theme in style.theme_names():
            style.theme_use(theme)
            break

    App(root)
    root.mainloop()


if __name__ == '__main__':
    main()
